# Document Management Views
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib.admin.views.decorators import staff_member_required
from django.db import models
from django.utils import timezone
from django.conf import settings
import mimetypes
import os
from datetime import datetime
from openpyxl import load_workbook

from .models import (
    Applicant, Programs, DocumentCategory, ManualDocument, PolicyDocument, 
    Learner_Profile, TrainerProfile, LMSTC_Documents
)


@csrf_exempt
@require_POST
def upload_document(request):
    """Handle document upload for different document types"""
    try:
        document_type = request.POST.get('document_type')
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        version = request.POST.get('version', '1.0')
        document_file = request.FILES.get('document_file')
        
        if not document_type or not title or not document_file:
            return JsonResponse({'success': False, 'error': 'Missing required fields'})
        
        # Validate file size (10MB limit)
        if document_file.size > 10 * 1024 * 1024:
            return JsonResponse({'success': False, 'error': 'File size must be less than 10MB'})
        
        # Validate file type
        allowed_types = ['application/pdf', 'application/msword', 
                        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                        'image/jpeg', 'image/png', 'image/jpg']
        
        file_type = mimetypes.guess_type(document_file.name)[0]
        if file_type not in allowed_types:
            return JsonResponse({'success': False, 'error': 'Invalid file type. Only PDF, DOC, DOCX, JPG, PNG files are allowed'})
        
        if document_type == 'applicant_profile':
            # Handle Applicant Profile upload - save to Learner_Profile
            applicant_id = request.POST.get('applicant_id')
            if not applicant_id:
                return JsonResponse({'success': False, 'error': 'Applicant selection is required for profile documents'})
            
            try:
                applicant = Applicant.objects.get(id=applicant_id)
                learner_profile, created = Learner_Profile.objects.get_or_create(
                    user=applicant,
                    defaults={
                        'last_name': applicant.last_name or '',
                        'first_name': applicant.first_name or '',
                        'email': applicant.email or '',
                    }
                )
                
                # Save the document file to the learner profile
                if 'id_picture' in document_file.name.lower() or 'photo' in document_file.name.lower():
                    learner_profile.id_picture = document_file
                elif 'signature' in document_file.name.lower():
                    learner_profile.applicant_signature = document_file
                
                learner_profile.save()
                
                return JsonResponse({
                    'success': True, 
                    'message': f'Applicant profile document uploaded successfully for {applicant.get_full_name() or applicant.username}'
                })
                
            except Applicant.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Selected applicant not found'})
        
        elif document_type == 'manual':
            # Handle Manual document upload
            category_id = request.POST.get('category_id')
            program_id = request.POST.get('program_id') or None
            batch = request.POST.get('batch') or None
            semester = request.POST.get('semester') or None
            
            if not category_id:
                return JsonResponse({'success': False, 'error': 'Category selection is required for manual documents'})
            
            try:
                category = DocumentCategory.objects.get(id=category_id, category_type='manual')
                program = Programs.objects.get(id=program_id) if program_id else None
                
                manual_doc = ManualDocument.objects.create(
                    title=title,
                    description=description,
                    document_file=document_file,
                    category=category,
                    program=program,
                    batch=batch,
                    semester=semester,
                    version=version,
                    uploaded_by=request.user
                )
                
                return JsonResponse({
                    'success': True, 
                    'message': f'Manual document "{title}" uploaded successfully'
                })
                
            except DocumentCategory.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Selected category not found'})
            except Programs.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Selected program not found'})
        
        elif document_type == 'policy':
            # Handle Policy document upload
            category_id = request.POST.get('category_id')
            program_id = request.POST.get('program_id') or None
            batch = request.POST.get('batch') or None
            semester = request.POST.get('semester') or None
            effective_date = request.POST.get('effective_date') or None
            expiry_date = request.POST.get('expiry_date') or None
            
            if not category_id:
                return JsonResponse({'success': False, 'error': 'Category selection is required for policy documents'})
            
            try:
                category = DocumentCategory.objects.get(id=category_id, category_type='policy')
                program = Programs.objects.get(id=program_id) if program_id else None
                
                policy_doc = PolicyDocument.objects.create(
                    title=title,
                    description=description,
                    document_file=document_file,
                    category=category,
                    program=program,
                    batch=batch,
                    semester=semester,
                    version=version,
                    uploaded_by=request.user,
                    effective_date=effective_date,
                    expiry_date=expiry_date
                )
                
                return JsonResponse({
                    'success': True, 
                    'message': f'Policy document "{title}" uploaded successfully'
                })
                
            except DocumentCategory.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Selected category not found'})
            except Programs.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Selected program not found'})
        
        else:
            return JsonResponse({'success': False, 'error': 'Invalid document type'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Upload failed: {str(e)}'})


def search_documents(request):
    """Search and filter documents based on criteria"""
    from .models import LMSTC_Documents, ApprovedApplicant
    
    try:
        # Get search parameters
        query = request.GET.get('q', '')
        document_type = request.GET.get('document_type', '')
        batch = request.GET.get('batch', '')
        semester = request.GET.get('semester', '')
        program_id = request.GET.get('program_id', '')
        trainor_id = request.GET.get('trainor_id', '')
        year = request.GET.get('year', '')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        
        results = []
        
        # Search Applicant Profiles (Learner_Profile)
        if not document_type or document_type == 'applicant_profile':
            learner_profiles = Learner_Profile.objects.select_related('user').all()
            
            if query:
                learner_profiles = learner_profiles.filter(
                    models.Q(first_name__icontains=query) |
                    models.Q(last_name__icontains=query) |
                    models.Q(user__username__icontains=query) |
                    models.Q(email__icontains=query)
                )
            
            for profile in learner_profiles:
                # Get BatchCycle info if available
                batch_display = 'N/A'
                semester_display = 'N/A'
                program_display = 'N/A'
                
                if profile.user:
                    try:
                        approved_app = ApprovedApplicant.objects.select_related('program').get(
                            applicant=profile.user,
                            status='active'
                        )
                        if approved_app.batch_number:
                            batch_display = f"Batch {approved_app.batch_number}"
                        if approved_app.enrollment_semester:
                            semester_display = f"Semester {approved_app.enrollment_semester}"
                        if approved_app.program:
                            program_display = approved_app.program.program_name
                    except ApprovedApplicant.DoesNotExist:
                        pass
                
                results.append({
                    'id': f'LP-{profile.id}',
                    'profile_id': profile.id,
                    'title': f"{profile.first_name} {profile.last_name} - Profile",
                    'type': 'Applicant Profile',
                    'content': f"Email: {profile.email}, Entry Date: {profile.entry_date or 'N/A'}",
                    'batch': batch_display,
                    'semester': semester_display,
                    'program': program_display,
                    'trainor': 'N/A',
                    'created_at': profile.entry_date,
                    'file_type': 'Learner Profile',
                    'source': 'learner_profile'
                })
            
            # Also search LMSTC_Documents with type='applicant_profile'
            lmstc_profiles = LMSTC_Documents.objects.select_related(
                'learner_profile', 'applicant', 'uploaded_by', 'program'
            ).filter(document_type='applicant_profile', status='active')
            
            if query:
                lmstc_profiles = lmstc_profiles.filter(
                    models.Q(document_name__icontains=query) |
                    models.Q(applicant__username__icontains=query) |
                    models.Q(applicant__first_name__icontains=query) |
                    models.Q(applicant__last_name__icontains=query) |
                    models.Q(learner_profile__first_name__icontains=query) |
                    models.Q(learner_profile__last_name__icontains=query)
                )
            
            # Apply year filter
            if year:
                lmstc_profiles = lmstc_profiles.filter(uploaded_at__year=year)
            
            if batch:
                lmstc_profiles = lmstc_profiles.filter(batch=batch)
            
            if program_id:
                lmstc_profiles = lmstc_profiles.filter(program_id=program_id)
            
            for doc in lmstc_profiles:
                # Determine display name
                display_name = doc.document_name
                if doc.learner_profile:
                    display_name = f"{doc.learner_profile.first_name} {doc.learner_profile.last_name} - {doc.document_name}"
                elif doc.applicant:
                    display_name = f"{doc.applicant.get_full_name() or doc.applicant.username} - {doc.document_name}"
                
                # Get BatchCycle info
                batch_display = 'N/A'
                semester_display = 'N/A'
                program_display = doc.program.program_name if doc.program else 'N/A'
                
                if doc.applicant:
                    try:
                        approved_app = ApprovedApplicant.objects.select_related('program').get(
                            applicant=doc.applicant,
                            status='active'
                        )
                        if approved_app.batch_number:
                            batch_display = f"Batch {approved_app.batch_number}"
                        if approved_app.enrollment_semester:
                            semester_display = f"Semester {approved_app.enrollment_semester}"
                        if approved_app.program:
                            program_display = approved_app.program.program_name
                    except ApprovedApplicant.DoesNotExist:
                        pass
                
                # Fallback to doc.batch if no ApprovedApplicant
                if batch_display == 'N/A' and doc.batch:
                    if '_' in doc.batch:
                        batch_num = doc.batch.split('_')[1]
                        batch_display = f"Batch {batch_num}"
                    else:
                        batch_display = f"Batch {doc.batch}"
                
                results.append({
                    'id': f'LMSTC-{doc.id}',
                    'document_id': doc.id,
                    'title': display_name,
                    'type': 'Applicant Profile',
                    'content': f"Uploaded: {doc.uploaded_at.strftime('%Y-%m-%d')}, By: {doc.uploaded_by.get_full_name() or doc.uploaded_by.username}",
                    'batch': batch_display,
                    'semester': semester_display,
                    'program': program_display,
                    'trainor': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                    'created_at': doc.uploaded_at,
                    'file_type': 'Uploaded Document',
                    'file_url': doc.document_file.url if doc.document_file else None,
                    'source': 'lmstc_document'
                })
        
        # Search Manual Documents
        if not document_type or document_type == 'manual':
            manual_docs = ManualDocument.objects.select_related('category', 'program', 'uploaded_by').filter(status='active')
            
            if query:
                manual_docs = manual_docs.filter(
                    models.Q(title__icontains=query) |
                    models.Q(description__icontains=query) |
                    models.Q(category__name__icontains=query)
                )
            
            if batch:
                manual_docs = manual_docs.filter(batch=batch)
            
            if semester:
                manual_docs = manual_docs.filter(semester=semester)
            
            if program_id:
                manual_docs = manual_docs.filter(program_id=program_id)
            
            for doc in manual_docs:
                results.append({
                    'id': doc.id,
                    'title': doc.title,
                    'type': 'Manual',
                    'content': f"Category: {doc.category.name}, Description: {doc.description or 'No description'}",
                    'batch': doc.get_batch_display() if doc.batch else 'All Batches',
                    'semester': doc.get_semester_display() if doc.semester else 'All Semesters',
                    'program': doc.program.program_name if doc.program else 'All Programs',
                    'trainor': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                    'created_at': doc.created_at,
                    'file_type': 'Manual'
                })
        
        # Search LMSTC Documents with type='modules_manuals' (bulk uploaded modules/manuals)
        if not document_type or document_type == 'modules_manuals':
            lmstc_manuals = LMSTC_Documents.objects.select_related(
                'program', 'uploaded_by'
            ).filter(document_type='modules_manuals', status='active')
            
            if query:
                lmstc_manuals = lmstc_manuals.filter(
                    models.Q(document_name__icontains=query) |
                    models.Q(description__icontains=query)
                )
            
            # Apply year filter
            if year:
                lmstc_manuals = lmstc_manuals.filter(uploaded_at__year=year)
            
            if batch:
                # Handle batch format conversion
                normalized = f"batch_{batch}" if not str(batch).startswith('batch_') else batch
                lmstc_manuals = lmstc_manuals.filter(batch__in=[batch, normalized])
            
            if program_id:
                lmstc_manuals = lmstc_manuals.filter(program_id=program_id)
            
            for doc in lmstc_manuals:
                batch_display = 'N/A'
                if doc.batch:
                    if '_' in doc.batch:
                        batch_num = doc.batch.split('_')[1]
                        batch_display = f"Batch {batch_num}"
                    else:
                        batch_display = f"Batch {doc.batch}"
                
                results.append({
                    'id': f'LMSTC-{doc.id}',
                    'document_id': doc.id,
                    'title': doc.document_name,
                    'type': 'Modules and Manuals',
                    'content': f"Description: {doc.description or 'No description'}",
                    'batch': batch_display,
                    'semester': 'N/A',
                    'program': doc.program.program_name if doc.program else 'N/A',
                    'trainor': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                    'created_at': doc.uploaded_at,
                    'file_type': 'Modules and Manuals',
                    'file_url': doc.document_file.url if (doc.document_file and hasattr(doc.document_file, 'name') and doc.document_file.name) else None,
                    'source': 'lmstc_document'
                })
        
        # Search LMSTC Documents with type='policy_guidelines' (bulk uploaded policies)
        if not document_type or document_type == 'policy_guidelines':
            lmstc_policies = LMSTC_Documents.objects.select_related(
                'program', 'uploaded_by'
            ).filter(document_type='policy_guidelines', status='active')
            
            if query:
                lmstc_policies = lmstc_policies.filter(
                    models.Q(document_name__icontains=query) |
                    models.Q(description__icontains=query)
                )
            
            # Apply year filter
            if year:
                lmstc_policies = lmstc_policies.filter(uploaded_at__year=year)
            
            if batch:
                # Handle batch format conversion
                normalized = f"batch_{batch}" if not str(batch).startswith('batch_') else batch
                lmstc_policies = lmstc_policies.filter(batch__in=[batch, normalized])
            
            if program_id:
                lmstc_policies = lmstc_policies.filter(program_id=program_id)
            
            for doc in lmstc_policies:
                batch_display = 'N/A'
                if doc.batch:
                    if '_' in doc.batch:
                        batch_num = doc.batch.split('_')[1]
                        batch_display = f"Batch {batch_num}"
                    else:
                        batch_display = f"Batch {doc.batch}"
                
                results.append({
                    'id': f'LMSTC-{doc.id}',
                    'document_id': doc.id,
                    'title': doc.document_name,
                    'type': 'Policy and Guidelines',
                    'content': f"Description: {doc.description or 'No description'}",
                    'batch': batch_display,
                    'semester': 'N/A',
                    'program': doc.program.program_name if doc.program else 'N/A',
                    'trainor': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                    'created_at': doc.uploaded_at,
                    'file_type': 'Policy and Guidelines',
                    'file_url': doc.document_file.url if (doc.document_file and hasattr(doc.document_file, 'name') and doc.document_file.name) else None,
                    'source': 'lmstc_document'
                })
        
        # Search Policy Documents (from PolicyDocument model - legacy)
        if not document_type or document_type == 'policy':
            policy_docs = PolicyDocument.objects.select_related('category', 'program', 'uploaded_by').filter(status='active')
            
            if query:
                policy_docs = policy_docs.filter(
                    models.Q(title__icontains=query) |
                    models.Q(description__icontains=query) |
                    models.Q(category__name__icontains=query)
                )
            
            if batch:
                policy_docs = policy_docs.filter(batch=batch)
            
            if semester:
                policy_docs = policy_docs.filter(semester=semester)
            
            if program_id:
                policy_docs = policy_docs.filter(program_id=program_id)
            
            for doc in policy_docs:
                results.append({
                    'id': doc.id,
                    'title': doc.title,
                    'type': 'Policy',
                    'content': f"Category: {doc.category.name}, Description: {doc.description or 'No description'}",
                    'batch': doc.get_batch_display() if doc.batch else 'All Batches',
                    'semester': doc.get_semester_display() if doc.semester else 'All Semesters',
                    'program': doc.program.program_name if doc.program else 'All Programs',
                    'trainor': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                    'created_at': doc.created_at,
                    'file_type': 'Policy'
                })
        
        # Sort results by creation date (newest first)
        results.sort(key=lambda x: x['created_at'] or timezone.now(), reverse=True)
        
        # Pagination
        total_results = len(results)
        start_index = (page - 1) * per_page
        end_index = start_index + per_page
        paginated_results = results[start_index:end_index]
        
        return JsonResponse({
            'success': True,
            'results': paginated_results,
            'total': total_results,
            'page': page,
            'per_page': per_page,
            'total_pages': (total_results + per_page - 1) // per_page
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Search failed: {str(e)}'})


def get_document_categories(request):
    """Get document categories based on type"""
    try:
        category_type = request.GET.get('type', '')
        
        if category_type not in ['manual', 'policy']:
            return JsonResponse({'success': False, 'error': 'Invalid category type'})
        
        categories = DocumentCategory.objects.filter(category_type=category_type).values('id', 'name', 'description')
        
        return JsonResponse({
            'success': True,
            'categories': list(categories)
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Failed to get categories: {str(e)}'})


@csrf_exempt
@require_POST
def bulk_upload_lmstc_documents(request):
    """Handle bulk upload of LMSTC documents"""
    from .models import LMSTC_Documents
    import os
    
    try:
        # Get form data
        document_type = request.POST.get('document_type')
        bulk_batch = request.POST.get('bulk_batch', '')
        bulk_program = request.POST.get('bulk_program', '')
        bulk_description = request.POST.get('bulk_description', '')
        
        if not document_type:
            return JsonResponse({'success': False, 'error': 'Document type is required'})
        
        if document_type not in ['applicant_profile', 'policy_guidelines', 'modules_manuals']:
            return JsonResponse({'success': False, 'error': 'Invalid document type'})
        
        # Check if user is authenticated
        if not request.user.is_authenticated:
            return JsonResponse({'success': False, 'error': 'User must be authenticated'})
        
        uploaded_files = []
        failed_files = []
        
        # Get all uploaded files
        files = request.FILES
        
        # Process each file
        for key in files:
            if key.startswith('file_'):
                file = files[key]
                
                # Validate file size (10MB limit)
                if file.size > 10 * 1024 * 1024:
                    failed_files.append({
                        'name': file.name,
                        'reason': 'File size exceeds 10MB'
                    })
                    continue
                
                # Validate file type
                allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx']
                file_ext = os.path.splitext(file.name)[1].lower()
                
                if file_ext not in allowed_extensions:
                    failed_files.append({
                        'name': file.name,
                        'reason': 'Invalid file type. Only PDF, DOC, DOCX, XLS, XLSX are allowed'
                    })
                    continue
                
                # Get individual file settings
                file_index = key.replace('file_', '')
                file_batch = request.POST.get(f'file_{file_index}_batch', bulk_batch)
                file_program = request.POST.get(f'file_{file_index}_program', bulk_program)
                
                try:
                    # Get program instance if provided
                    program_instance = None
                    if file_program:
                        try:
                            program_instance = Programs.objects.get(id=file_program)
                        except Programs.DoesNotExist:
                            pass
                    
                    # Get MIME type
                    mime_type = mimetypes.guess_type(file.name)[0] or 'application/octet-stream'
                    
                    # Create document record
                    document = LMSTC_Documents.objects.create(
                        document_name=file.name,
                        document_type=document_type,
                        document_file=file,
                        description=bulk_description,
                        batch=file_batch if file_batch else None,
                        program=program_instance,
                        uploaded_by=request.user,
                        file_size=file.size,
                        mime_type=mime_type,
                        status='active'
                    )
                    
                    uploaded_files.append({
                        'name': file.name,
                        'id': document.id,
                        'size': file.size
                    })
                    
                except Exception as e:
                    failed_files.append({
                        'name': file.name,
                        'reason': str(e)
                    })
        
        # Prepare response
        response_data = {
            'success': True,
            'message': f'Successfully uploaded {len(uploaded_files)} file(s)',
            'uploaded_count': len(uploaded_files),
            'failed_count': len(failed_files),
            'uploaded_files': uploaded_files,
            'failed_files': failed_files
        }
        
        if len(uploaded_files) == 0:
            response_data['success'] = False
            response_data['error'] = 'No files were uploaded successfully'
        
        return JsonResponse(response_data)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Bulk upload failed: {str(e)}'
        })


def get_bulk_download_documents(request):
    """Get combined Learner_Profile and LMSTC_Documents data for bulk download with BatchCycle info"""
    from .models import LMSTC_Documents, Learner_Profile, ApprovedApplicant, ApprovedWalkIn
    from django.db.models import Q
    
    try:
        # Get filter parameters
        search_query = request.GET.get('search', '')
        document_type = request.GET.get('document_type', '')
        year = request.GET.get('year', '')
        batch = request.GET.get('batch', '')
        program_id = request.GET.get('program', '')
        
        results = []
        
        # Fetch LMSTC Documents
        if not document_type or document_type in ['applicant_profile', 'policy_guidelines', 'modules_manuals']:
            lmstc_docs = LMSTC_Documents.objects.select_related(
                'learner_profile', 'applicant', 'uploaded_by'
            ).filter(status='active')
            
            # Apply document type filter
            if document_type:
                lmstc_docs = lmstc_docs.filter(document_type=document_type)
            
            # Apply year filter
            if year:
                lmstc_docs = lmstc_docs.filter(uploaded_at__year=year)
            
            # Apply batch filter
            if batch:
                # UI sends '1', '2', '3' while model stores 'batch_1', etc.
                # Match either format to be safe
                normalized = f"batch_{batch}" if not str(batch).startswith('batch_') else batch
                lmstc_docs = lmstc_docs.filter(batch__in=[batch, normalized])
            
            # Apply search filter
            if search_query:
                lmstc_docs = lmstc_docs.filter(
                    Q(document_name__icontains=search_query) |
                    Q(applicant__username__icontains=search_query) |
                    Q(applicant__first_name__icontains=search_query) |
                    Q(applicant__last_name__icontains=search_query) |
                    Q(learner_profile__first_name__icontains=search_query) |
                    Q(learner_profile__last_name__icontains=search_query)
                )
            
            for doc in lmstc_docs:
                # Determine owner name
                owner_name = 'Unknown'
                if doc.learner_profile:
                    owner_name = f"{doc.learner_profile.first_name} {doc.learner_profile.last_name}"
                elif doc.applicant:
                    owner_name = doc.applicant.get_full_name() or doc.applicant.username
                
                # Get BatchCycle info from ApprovedApplicant
                batch_display = 'N/A'
                year_display = doc.uploaded_at.year if doc.uploaded_at else None
                semester_display = 'N/A'
                
                # Prefer explicit applicant, otherwise fall back to learner_profile.user
                applicant_user = doc.applicant
                if not applicant_user and doc.learner_profile and getattr(doc.learner_profile, 'user', None):
                    applicant_user = doc.learner_profile.user
                
                if applicant_user:
                    try:
                        # Use the most recent ApprovedApplicant regardless of status
                        approved_app = ApprovedApplicant.objects.filter(
                            applicant=applicant_user
                        ).order_by('-approved_at').first()
                        if approved_app:
                            if approved_app.batch_number:
                                batch_display = f"Batch {approved_app.batch_number}"
                            if approved_app.enrollment_year:
                                year_display = approved_app.enrollment_year
                            if approved_app.enrollment_semester:
                                semester_display = f"Semester {approved_app.enrollment_semester}"
                        else:
                            # Fallback to ApprovedWalkIn if present
                            walkin_app = ApprovedWalkIn.objects.filter(
                                applicant=applicant_user
                            ).order_by('-approved_at').first()
                            if walkin_app:
                                if walkin_app.batch_number:
                                    batch_display = f"Batch {walkin_app.batch_number}"
                                if walkin_app.enrollment_year:
                                    year_display = walkin_app.enrollment_year
                    except Exception:
                        pass
                
                # Fallback to doc.batch if no ApprovedApplicant
                if batch_display == 'N/A' and doc.batch:
                    if '_' in doc.batch:
                        batch_num = doc.batch.split('_')[1]
                        batch_display = f"Batch {batch_num}"
                    else:
                        batch_display = f"Batch {doc.batch}"
                
                # Get file URL - check if file exists and has a name
                file_url = None
                if doc.document_file and hasattr(doc.document_file, 'name') and doc.document_file.name:
                    try:
                        file_url = doc.document_file.url
                    except (ValueError, AttributeError):
                        file_url = None
                
                results.append({
                    'id': f'LMSTC-{doc.id}',
                    'document_id': doc.id,
                    'document_name': doc.document_name,
                    'document_type': doc.get_document_type_display(),
                    'document_type_value': doc.document_type,
                    'owner': owner_name,
                    'batch': batch_display,
                    'semester': semester_display,
                    'uploaded_at': doc.uploaded_at.strftime('%Y-%m-%d') if doc.uploaded_at else 'N/A',
                    'year': year_display,
                    'status': 'Active',
                    'source': 'lmstc_document',
                    'file_url': file_url
                })
        
        # Fetch Learner Profiles (if document type is applicant_profile or not specified)
        if not document_type or document_type == 'applicant_profile':
            learner_profiles = Learner_Profile.objects.select_related('user').all()
            
            # Apply search filter
            if search_query:
                learner_profiles = learner_profiles.filter(
                    Q(first_name__icontains=search_query) |
                    Q(last_name__icontains=search_query) |
                    Q(user__username__icontains=search_query) |
                    Q(email__icontains=search_query)
                )
            
            for profile in learner_profiles:
                # Determine the year from entry_date or date_accomplished
                year_value = None
                date_str = 'N/A'
                if profile.entry_date:
                    year_value = profile.entry_date.year
                    date_str = profile.entry_date.strftime('%Y-%m-%d')
                elif profile.date_accomplished:
                    year_value = profile.date_accomplished.year
                    date_str = profile.date_accomplished.strftime('%Y-%m-%d')
                
                # Get BatchCycle info from ApprovedApplicant
                batch_display = 'N/A'
                semester_display = 'N/A'
                
                has_approval_record = False
                
                if profile.user:
                    try:
                        # Use the most recent ApprovedApplicant regardless of status
                        approved_app = ApprovedApplicant.objects.select_related('program').filter(
                            applicant=profile.user
                        ).order_by('-approved_at').first()
                        if approved_app:
                            has_approval_record = True
                            # Apply filters based on ApprovedApplicant
                            if batch and approved_app.batch_number != batch:
                                continue
                            if year and approved_app.enrollment_year and approved_app.enrollment_year != int(year):
                                continue
                            if program_id and approved_app.program_id != int(program_id):
                                continue
                            
                            # Set batch and semester display
                            if approved_app.batch_number:
                                batch_display = f"Batch {approved_app.batch_number}"
                            if approved_app.enrollment_year:
                                year_value = approved_app.enrollment_year
                            if approved_app.enrollment_semester:
                                semester_display = f"Semester {approved_app.enrollment_semester}"
                        else:
                            # Try walk-in approvals
                            walkin_app = ApprovedWalkIn.objects.filter(
                                applicant=profile.user
                            ).order_by('-approved_at').first()
                            if walkin_app:
                                has_approval_record = True
                                if batch and walkin_app.batch_number != batch:
                                    continue
                                if year and walkin_app.enrollment_year and walkin_app.enrollment_year != int(year):
                                    continue
                                # Set display from walk-in
                                if walkin_app.batch_number:
                                    batch_display = f"Batch {walkin_app.batch_number}"
                                if walkin_app.enrollment_year:
                                    year_value = walkin_app.enrollment_year
                    except Exception:
                        # On any error, if filters are applied and we cannot confidently match,
                        # fall back to entry_date/year_value for year filtering
                        if year:
                            try:
                                if not year_value or year_value != int(year):
                                    continue
                            except ValueError:
                                continue
                        if batch or program_id:
                            continue
                else:
                    # No user - if filters are applied use entry_date/year_value for year filtering,
                    # but skip when batch/program filters are present (no way to resolve them)
                    if batch or program_id:
                        continue
                    if year:
                        try:
                            if not year_value or year_value != int(year):
                                continue
                        except ValueError:
                            continue
                
                # If we have a year filter and no approval/walk-in record, fall back to entry_date/year_value
                if year and not has_approval_record:
                    try:
                        if not year_value or year_value != int(year):
                            continue
                    except ValueError:
                        continue
                
                results.append({
                    'id': f'LP-{profile.id}',
                    'document_id': profile.id,
                    'document_name': f"{profile.first_name} {profile.last_name} - Profile",
                    'document_type': 'Learner Profile',
                    'document_type_value': 'learner_profile',
                    'owner': f"{profile.first_name} {profile.last_name}",
                    'batch': batch_display,
                    'semester': semester_display,
                    'uploaded_at': date_str,
                    'year': year_value,
                    'status': 'Active',
                    'source': 'learner_profile',
                    'file_url': None  # Learner profiles don't have direct file URLs
                })
        
        return JsonResponse({
            'success': True,
            'documents': results,
            'total_count': len(results)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to fetch documents: {str(e)}'
        })


def get_lmstc_documents(request):
    """Get LMSTC documents with filtering"""
    from .models import LMSTC_Documents, ApprovedApplicant, ApprovedWalkIn, TrainerProfile
    from django.db.models import Q
    
    try:
        document_type = request.GET.get('document_type', '')
        batch = request.GET.get('batch', '')
        program_id = request.GET.get('program_id', '')
        status = request.GET.get('status', 'active')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        
        # Base query
        documents = LMSTC_Documents.objects.select_related('program', 'applicant', 'uploaded_by').filter(status=status)
        
        # Apply filters
        if document_type:
            documents = documents.filter(document_type=document_type)
        
        if batch:
            documents = documents.filter(batch=batch)
        
        if program_id:
            documents = documents.filter(program_id=program_id)
        
        # Filter by user's approved programs (only for non-staff users)
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            # Get user's approved programs (from both ApprovedApplicant and ApprovedWalkIn)
            approved_program_ids = set()
            
            # Get from ApprovedApplicant
            approved_applicants = ApprovedApplicant.objects.filter(applicant=request.user).values_list('program_id', flat=True)
            approved_program_ids.update(approved_applicants)
            
            # Get from ApprovedWalkIn
            approved_walkins = ApprovedWalkIn.objects.filter(applicant=request.user).values_list('program_id', flat=True)
            approved_program_ids.update(approved_walkins)
            
            if approved_program_ids:
                # Filter documents based on program assignment
                # Documents should be visible if:
                # 1. Document has a program AND user is approved for that program
                # 2. Document has no program BUT uploaded_by is a trainer AND user is approved for any of trainer's programs
                
                # Build query for documents with explicit program matching user's approved programs
                program_filter = Q(program_id__in=approved_program_ids)
                
                # Build query for documents without program but uploaded by trainers
                # Get all trainers who have programs that match user's approved programs
                trainers_with_matching_programs = TrainerProfile.objects.filter(
                    programs__id__in=approved_program_ids
                ).values_list('user_id', flat=True).distinct()
                
                # Documents without program uploaded by trainers with matching programs
                trainer_filter = Q(program__isnull=True) & Q(uploaded_by_id__in=trainers_with_matching_programs)
                
                # Combine both conditions
                documents = documents.filter(program_filter | trainer_filter)
            else:
                # User has no approved programs, show no documents
                documents = documents.none()
        
        # Order by upload date (newest first)
        documents = documents.order_by('-uploaded_at')
        
        # Count total
        total_count = documents.count()
        
        # Pagination
        start_index = (page - 1) * per_page
        end_index = start_index + per_page
        documents = documents[start_index:end_index]
        
        # Format results
        results = []
        for doc in documents:
            # Get file URL - check if file exists and has a name
            file_url = None
            if doc.document_file and hasattr(doc.document_file, 'name') and doc.document_file.name:
                try:
                    file_url = doc.document_file.url
                except (ValueError, AttributeError):
                    file_url = None
            
            results.append({
                'id': doc.id,
                'document_name': doc.document_name,
                'document_type': doc.get_document_type_display(),
                'document_type_value': doc.document_type,
                'batch': doc.get_batch_display() if doc.batch else 'N/A',
                'program': doc.program.program_name if doc.program else 'N/A',
                'applicant': f"{doc.applicant.first_name} {doc.applicant.last_name}" if doc.applicant else 'N/A',
                'uploaded_by': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                'uploaded_at': doc.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                'file_url': file_url,
                'file_size': doc.get_file_size_display(),
                'description': doc.description or ''
            })
        
        return JsonResponse({
            'success': True,
            'documents': results,
            'total_count': total_count,
            'page': page,
            'per_page': per_page
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to fetch documents: {str(e)}'
        })


@csrf_exempt
@require_POST
def archive_document(request):
    """Archive a single document (soft delete)"""
    from .models import LMSTC_Documents
    
    try:
        document_id = request.POST.get('document_id')
        document_type = request.POST.get('document_type', 'lmstc')  # lmstc, manual, policy
        
        if not document_id:
            return JsonResponse({'success': False, 'error': 'Document ID is required'})
        
        # Archive based on document type
        if document_type == 'lmstc':
            try:
                document = LMSTC_Documents.objects.get(id=document_id)
                document.status = 'archived'
                document.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Document "{document.document_name}" has been archived successfully'
                })
            except LMSTC_Documents.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Document not found'})
        
        elif document_type == 'manual':
            try:
                document = ManualDocument.objects.get(id=document_id)
                document.status = 'archived'
                document.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Manual "{document.title}" has been archived successfully'
                })
            except ManualDocument.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Manual document not found'})
        
        elif document_type == 'policy':
            try:
                document = PolicyDocument.objects.get(id=document_id)
                document.status = 'archived'
                document.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Policy "{document.title}" has been archived successfully'
                })
            except PolicyDocument.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Policy document not found'})
        
        else:
            return JsonResponse({'success': False, 'error': 'Invalid document type'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Archive failed: {str(e)}'})


@csrf_exempt
@require_POST
def delete_document(request):
    """Permanently delete a document (admin only)"""
    from .models import LMSTC_Documents
    import os
    
    try:
        # Check if user is staff/admin
        if not request.user.is_staff:
            return JsonResponse({'success': False, 'error': 'Permission denied. Only administrators can permanently delete documents.'})
        
        document_id = request.POST.get('document_id')
        document_type = request.POST.get('document_type', 'lmstc')
        
        if not document_id:
            return JsonResponse({'success': False, 'error': 'Document ID is required'})
        
        # Delete based on document type
        if document_type == 'lmstc':
            try:
                document = LMSTC_Documents.objects.get(id=document_id)
                document_name = document.document_name
                
                # Delete the physical file if it exists
                if document.document_file:
                    try:
                        if os.path.isfile(document.document_file.path):
                            os.remove(document.document_file.path)
                    except Exception as e:
                        pass  # Continue even if file deletion fails
                
                document.delete()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Document "{document_name}" has been permanently deleted'
                })
            except LMSTC_Documents.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Document not found'})
        
        elif document_type == 'manual':
            try:
                document = ManualDocument.objects.get(id=document_id)
                document_name = document.title
                
                # Delete the physical file
                if document.document_file:
                    try:
                        if os.path.isfile(document.document_file.path):
                            os.remove(document.document_file.path)
                    except Exception as e:
                        pass
                
                document.delete()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Manual "{document_name}" has been permanently deleted'
                })
            except ManualDocument.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Manual document not found'})
        
        elif document_type == 'policy':
            try:
                document = PolicyDocument.objects.get(id=document_id)
                document_name = document.title
                
                # Delete the physical file
                if document.document_file:
                    try:
                        if os.path.isfile(document.document_file.path):
                            os.remove(document.document_file.path)
                    except Exception as e:
                        pass
                
                document.delete()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Policy "{document_name}" has been permanently deleted'
                })
            except PolicyDocument.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Policy document not found'})
        
        else:
            return JsonResponse({'success': False, 'error': 'Invalid document type'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Delete failed: {str(e)}'})


@csrf_exempt
@require_POST
def restore_document(request):
    """Restore an archived document"""
    from .models import LMSTC_Documents
    
    try:
        document_id = request.POST.get('document_id')
        document_type = request.POST.get('document_type', 'lmstc')
        
        if not document_id:
            return JsonResponse({'success': False, 'error': 'Document ID is required'})
        
        # Restore based on document type
        if document_type == 'lmstc':
            try:
                document = LMSTC_Documents.objects.get(id=document_id, status='archived')
                document.status = 'active'
                document.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Document "{document.document_name}" has been restored successfully'
                })
            except LMSTC_Documents.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Archived document not found'})
        
        elif document_type == 'manual':
            try:
                document = ManualDocument.objects.get(id=document_id, status='archived')
                document.status = 'active'
                document.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Manual "{document.title}" has been restored successfully'
                })
            except ManualDocument.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Archived manual not found'})
        
        elif document_type == 'policy':
            try:
                document = PolicyDocument.objects.get(id=document_id, status='archived')
                document.status = 'active'
                document.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Policy "{document.title}" has been restored successfully'
                })
            except PolicyDocument.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Archived policy not found'})
        
        else:
            return JsonResponse({'success': False, 'error': 'Invalid document type'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Restore failed: {str(e)}'})


@csrf_exempt
@require_POST
def bulk_archive_documents(request):
    """Archive multiple documents at once"""
    from .models import LMSTC_Documents
    import json
    
    try:
        document_ids = request.POST.get('document_ids')
        document_type = request.POST.get('document_type', 'lmstc')
        
        if not document_ids:
            return JsonResponse({'success': False, 'error': 'Document IDs are required'})
        
        # Parse JSON array of IDs
        try:
            ids_list = json.loads(document_ids) if isinstance(document_ids, str) else document_ids
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid document IDs format'})
        
        archived_count = 0
        failed_count = 0
        
        # Archive based on document type
        if document_type == 'lmstc':
            for doc_id in ids_list:
                try:
                    document = LMSTC_Documents.objects.get(id=doc_id)
                    document.status = 'archived'
                    document.save()
                    archived_count += 1
                except LMSTC_Documents.DoesNotExist:
                    failed_count += 1
        
        elif document_type == 'manual':
            for doc_id in ids_list:
                try:
                    document = ManualDocument.objects.get(id=doc_id)
                    document.status = 'archived'
                    document.save()
                    archived_count += 1
                except ManualDocument.DoesNotExist:
                    failed_count += 1
        
        elif document_type == 'policy':
            for doc_id in ids_list:
                try:
                    document = PolicyDocument.objects.get(id=doc_id)
                    document.status = 'archived'
                    document.save()
                    archived_count += 1
                except PolicyDocument.DoesNotExist:
                    failed_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully archived {archived_count} document(s)',
            'archived_count': archived_count,
            'failed_count': failed_count
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Bulk archive failed: {str(e)}'})


@csrf_exempt
@require_POST
def bulk_delete_documents(request):
    """Permanently delete multiple documents (admin only)"""
    from .models import LMSTC_Documents
    import json
    import os
    
    try:
        # Check if user is staff/admin
        if not request.user.is_staff:
            return JsonResponse({'success': False, 'error': 'Permission denied. Only administrators can permanently delete documents.'})
        
        document_ids = request.POST.get('document_ids')
        document_type = request.POST.get('document_type', 'lmstc')
        
        if not document_ids:
            return JsonResponse({'success': False, 'error': 'Document IDs are required'})
        
        # Parse JSON array of IDs
        try:
            ids_list = json.loads(document_ids) if isinstance(document_ids, str) else document_ids
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid document IDs format'})
        
        deleted_count = 0
        failed_count = 0
        
        # Delete based on document type
        if document_type == 'lmstc':
            for doc_id in ids_list:
                try:
                    document = LMSTC_Documents.objects.get(id=doc_id)
                    
                    # Delete the physical file
                    if document.document_file:
                        try:
                            if os.path.isfile(document.document_file.path):
                                os.remove(document.document_file.path)
                        except Exception:
                            pass
                    
                    document.delete()
                    deleted_count += 1
                except LMSTC_Documents.DoesNotExist:
                    failed_count += 1
        
        elif document_type == 'manual':
            for doc_id in ids_list:
                try:
                    document = ManualDocument.objects.get(id=doc_id)
                    
                    if document.document_file:
                        try:
                            if os.path.isfile(document.document_file.path):
                                os.remove(document.document_file.path)
                        except Exception:
                            pass
                    
                    document.delete()
                    deleted_count += 1
                except ManualDocument.DoesNotExist:
                    failed_count += 1
        
        elif document_type == 'policy':
            for doc_id in ids_list:
                try:
                    document = PolicyDocument.objects.get(id=doc_id)
                    
                    if document.document_file:
                        try:
                            if os.path.isfile(document.document_file.path):
                                os.remove(document.document_file.path)
                        except Exception:
                            pass
                    
                    document.delete()
                    deleted_count += 1
                except PolicyDocument.DoesNotExist:
                    failed_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully deleted {deleted_count} document(s)',
            'deleted_count': deleted_count,
            'failed_count': failed_count
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Bulk delete failed: {str(e)}'})


def get_archived_documents(request):
    """Get all archived documents with filtering"""
    from .models import LMSTC_Documents
    
    try:
        document_type = request.GET.get('document_type', '')
        search_query = request.GET.get('search', '')
        
        results = []
        
        # Get archived LMSTC documents
        if not document_type or document_type == 'lmstc':
            lmstc_docs = LMSTC_Documents.objects.filter(status='archived').select_related(
                'uploaded_by', 'applicant', 'program'
            ).order_by('-updated_at')
            
            if search_query:
                lmstc_docs = lmstc_docs.filter(
                    models.Q(document_name__icontains=search_query) |
                    models.Q(description__icontains=search_query)
                )
            
            for doc in lmstc_docs:
                results.append({
                    'id': doc.id,
                    'document_type': 'lmstc',
                    'name': doc.document_name,
                    'type_display': doc.get_document_type_display(),
                    'uploaded_by': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                    'uploaded_at': doc.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                    'archived_at': doc.updated_at.strftime('%Y-%m-%d %H:%M'),
                    'batch': doc.get_batch_display() if doc.batch else 'N/A',
                    'program': doc.program.program_name if doc.program else 'N/A'
                })
        
        # Get archived Manual documents
        if not document_type or document_type == 'manual':
            manual_docs = ManualDocument.objects.filter(status='archived').select_related(
                'uploaded_by', 'category', 'program'
            ).order_by('-updated_at')
            
            if search_query:
                manual_docs = manual_docs.filter(
                    models.Q(title__icontains=search_query) |
                    models.Q(description__icontains=search_query)
                )
            
            for doc in manual_docs:
                results.append({
                    'id': doc.id,
                    'document_type': 'manual',
                    'name': doc.title,
                    'type_display': f'Manual - {doc.category.name}',
                    'uploaded_by': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                    'uploaded_at': doc.created_at.strftime('%Y-%m-%d %H:%M'),
                    'archived_at': doc.updated_at.strftime('%Y-%m-%d %H:%M'),
                    'batch': doc.get_batch_display() if doc.batch else 'All Batches',
                    'program': doc.program.program_name if doc.program else 'All Programs'
                })
        
        # Get archived Policy documents
        if not document_type or document_type == 'policy':
            policy_docs = PolicyDocument.objects.filter(status='archived').select_related(
                'uploaded_by', 'category', 'program'
            ).order_by('-updated_at')
            
            if search_query:
                policy_docs = policy_docs.filter(
                    models.Q(title__icontains=search_query) |
                    models.Q(description__icontains=search_query)
                )
            
            for doc in policy_docs:
                results.append({
                    'id': doc.id,
                    'document_type': 'policy',
                    'name': doc.title,
                    'type_display': f'Policy - {doc.category.name}',
                    'uploaded_by': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                    'uploaded_at': doc.created_at.strftime('%Y-%m-%d %H:%M'),
                    'archived_at': doc.updated_at.strftime('%Y-%m-%d %H:%M'),
                    'batch': doc.get_batch_display() if doc.batch else 'All Batches',
                    'program': doc.program.program_name if doc.program else 'All Programs'
                })
        
        return JsonResponse({
            'success': True,
            'documents': results,
            'total_count': len(results)
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Failed to fetch archived documents: {str(e)}'})
        for doc in documents:
            results.append({
                'id': doc.id,
                'document_name': doc.document_name,
                'document_type': doc.get_document_type_display(),
                'document_type_value': doc.document_type,
                'batch': doc.get_batch_display() if doc.batch else 'N/A',
                'program': doc.program.program_name if doc.program else 'N/A',
                'uploaded_by': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                'uploaded_at': doc.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                'file_size': doc.get_file_size_display(),
                'status': doc.get_status_display(),
                'file_url': doc.document_file.url if doc.document_file else None,
                'description': doc.description or ''
            })
        
        return JsonResponse({
            'success': True,
            'documents': results,
            'total': total_count,
            'page': page,
            'per_page': per_page,
            'total_pages': (total_count + per_page - 1) // per_page
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to retrieve documents: {str(e)}'
        })


@csrf_exempt
@require_POST
def archive_document(request):
    """Archive a document (soft delete)"""
    from .models import DocumentAuditLog
    
    try:
        document_type = request.POST.get('document_type')
        document_id = request.POST.get('document_id')
        reason = request.POST.get('reason', '')
        
        if not document_type or not document_id:
            return JsonResponse({'success': False, 'error': 'Missing required fields'})
        
        document = None
        document_name = ''
        
        # Get the document based on type
        if document_type == 'manual':
            document = ManualDocument.objects.get(id=document_id, status='active')
            document_name = document.title
        elif document_type == 'policy':
            document = PolicyDocument.objects.get(id=document_id, status='active')
            document_name = document.title
        elif document_type == 'lmstc_document' or document_type == 'lmstc':
            from .models import LMSTC_Documents
            try:
                document = LMSTC_Documents.objects.get(id=document_id, status='active')
                document_name = document.document_name
            except LMSTC_Documents.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Document not found or already archived'})
        else:
            return JsonResponse({'success': False, 'error': 'Invalid document type'})
        
        # Archive the document
        document.status = 'archived'
        document.save()
        
        # Log the action
        DocumentAuditLog.objects.create(
            document_type=document_type,
            document_id=document_id,
            document_name=document_name,
            action='archive',
            performed_by=request.user,
            reason=reason,
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Document "{document_name}" archived successfully'
        })
    
    except (ManualDocument.DoesNotExist, PolicyDocument.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Document not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Archive failed: {str(e)}'})


@csrf_exempt
@require_POST
def restore_document(request):
    """Restore an archived document"""
    from .models import DocumentAuditLog
    
    try:
        document_type = request.POST.get('document_type')
        document_id = request.POST.get('document_id')
        
        if not document_type or not document_id:
            return JsonResponse({'success': False, 'error': 'Missing required fields'})
        
        document = None
        document_name = ''
        
        # Get the document based on type
        if document_type == 'manual':
            document = ManualDocument.objects.get(id=document_id)
            document_name = document.title
        elif document_type == 'policy':
            document = PolicyDocument.objects.get(id=document_id)
            document_name = document.title
        elif document_type == 'lmstc_document':
            from .models import LMSTC_Documents
            document = LMSTC_Documents.objects.get(id=document_id)
            document_name = document.document_name
        else:
            return JsonResponse({'success': False, 'error': 'Invalid document type'})
        
        # Restore the document
        document.status = 'active'
        document.save()
        
        # Log the action
        DocumentAuditLog.objects.create(
            document_type=document_type,
            document_id=document_id,
            document_name=document_name,
            action='restore',
            performed_by=request.user,
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Document "{document_name}" restored successfully'
        })
    
    except (ManualDocument.DoesNotExist, PolicyDocument.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Document not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Restore failed: {str(e)}'})


@csrf_exempt
@require_POST
def delete_document(request):
    """Permanently delete a document (Admin only)"""
    from .models import DocumentAuditLog
    import os
    
    try:
        # Check if user is staff/admin
        if not request.user.is_staff:
            return JsonResponse({'success': False, 'error': 'Permission denied. Admin access required.'})
        
        document_type = request.POST.get('document_type')
        document_id = request.POST.get('document_id')
        reason = request.POST.get('reason', '')
        
        if not document_type or not document_id:
            return JsonResponse({'success': False, 'error': 'Missing required fields'})
        
        if not reason:
            return JsonResponse({'success': False, 'error': 'Deletion reason is required'})
        
        document = None
        document_name = ''
        file_path = None
        
        # Get the document based on type
        if document_type == 'manual':
            document = ManualDocument.objects.get(id=document_id)
            document_name = document.title
            file_path = document.document_file.path if document.document_file else None
        elif document_type == 'policy':
            document = PolicyDocument.objects.get(id=document_id)
            document_name = document.title
            file_path = document.document_file.path if document.document_file else None
        elif document_type == 'lmstc_document':
            from .models import LMSTC_Documents
            document = LMSTC_Documents.objects.get(id=document_id)
            document_name = document.document_name
            file_path = document.document_file.path if document.document_file else None
        else:
            return JsonResponse({'success': False, 'error': 'Invalid document type'})
        
        # Log the action before deletion
        DocumentAuditLog.objects.create(
            document_type=document_type,
            document_id=document_id,
            document_name=document_name,
            action='delete',
            performed_by=request.user,
            reason=reason,
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        # Delete the physical file
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as e:
                print(f"Error deleting file: {e}")
        
        # Delete the database record
        document.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Document "{document_name}" permanently deleted'
        })
    
    except (ManualDocument.DoesNotExist, PolicyDocument.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Document not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Delete failed: {str(e)}'})


@csrf_exempt
@require_POST
def edit_document(request):
    """Edit document details"""
    from .models import DocumentAuditLog
    
    try:
        document_type = request.POST.get('document_type')
        document_id = request.POST.get('document_id')
        
        if not document_type or not document_id:
            return JsonResponse({'success': False, 'error': 'Missing required fields'})
        
        document = None
        document_name = ''
        previous_data = {}
        new_data = {}
        
        # Get the document based on type
        if document_type == 'manual':
            document = ManualDocument.objects.get(id=document_id)
            document_name = document.title
            previous_data = {
                'title': document.title,
                'description': document.description,
                'version': document.version
            }
            
            # Update fields if provided
            if request.POST.get('title'):
                document.title = request.POST.get('title')
            if request.POST.get('description'):
                document.description = request.POST.get('description')
            if request.POST.get('version'):
                document.version = request.POST.get('version')
            
            document.save()
            new_data = {
                'title': document.title,
                'description': document.description,
                'version': document.version
            }
            
        elif document_type == 'policy':
            document = PolicyDocument.objects.get(id=document_id)
            document_name = document.title
            previous_data = {
                'title': document.title,
                'description': document.description,
                'version': document.version
            }
            
            # Update fields if provided
            if request.POST.get('title'):
                document.title = request.POST.get('title')
            if request.POST.get('description'):
                document.description = request.POST.get('description')
            if request.POST.get('version'):
                document.version = request.POST.get('version')
            
            document.save()
            new_data = {
                'title': document.title,
                'description': document.description,
                'version': document.version
            }
            
        elif document_type == 'lmstc_document':
            from .models import LMSTC_Documents
            document = LMSTC_Documents.objects.get(id=document_id)
            document_name = document.document_name
            previous_data = {
                'document_name': document.document_name,
                'description': document.description
            }
            
            # Update fields if provided
            if request.POST.get('document_name'):
                document.document_name = request.POST.get('document_name')
            if request.POST.get('description'):
                document.description = request.POST.get('description')
            
            document.save()
            new_data = {
                'document_name': document.document_name,
                'description': document.description
            }
        else:
            return JsonResponse({'success': False, 'error': 'Invalid document type'})
        
        # Log the action
        DocumentAuditLog.objects.create(
            document_type=document_type,
            document_id=document_id,
            document_name=document_name,
            action='edit',
            performed_by=request.user,
            previous_data=previous_data,
            new_data=new_data,
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Document "{document_name}" updated successfully'
        })
    
    except (ManualDocument.DoesNotExist, PolicyDocument.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Document not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Edit failed: {str(e)}'})


def get_document_audit_logs(request):
    """Get audit logs for documents"""
    from .models import DocumentAuditLog
    
    try:
        document_type = request.GET.get('document_type', '')
        document_id = request.GET.get('document_id', '')
        action = request.GET.get('action', '')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        
        # Base query
        logs = DocumentAuditLog.objects.select_related('performed_by').all()
        
        # Apply filters
        if document_type:
            logs = logs.filter(document_type=document_type)
        
        if document_id:
            logs = logs.filter(document_id=document_id)
        
        if action:
            logs = logs.filter(action=action)
        
        # Count total
        total_count = logs.count()
        
        # Pagination
        start_index = (page - 1) * per_page
        end_index = start_index + per_page
        logs = logs[start_index:end_index]
        
        # Format results
        results = []
        for log in logs:
            results.append({
                'id': log.id,
                'document_type': log.document_type,
                'document_id': log.document_id,
                'document_name': log.document_name,
                'action': log.get_action_display(),
                'action_value': log.action,
                'performed_by': log.performed_by.get_full_name() or log.performed_by.username,
                'performed_at': log.performed_at.strftime('%Y-%m-%d %H:%M:%S'),
                'reason': log.reason or '',
                'ip_address': log.ip_address or '',
            })
        
        return JsonResponse({
            'success': True,
            'logs': results,
            'total': total_count,
            'page': page,
            'per_page': per_page,
            'total_pages': (total_count + per_page - 1) // per_page
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to retrieve audit logs: {str(e)}'
        })


def get_archived_documents(request):
    """Get all archived documents"""
    from .models import LMSTC_Documents
    
    try:
        document_type = request.GET.get('document_type', '')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        
        results = []
        
        # Get archived Manual Documents
        if not document_type or document_type == 'manual':
            manual_docs = ManualDocument.objects.select_related('category', 'program', 'uploaded_by').filter(status='archived')
            for doc in manual_docs:
                results.append({
                    'id': doc.id,
                    'document_type': 'manual',
                    'document_name': doc.title,
                    'description': doc.description or '',
                    'category': doc.category.name if doc.category else 'N/A',
                    'program': doc.program.program_name if doc.program else 'N/A',
                    'batch': doc.get_batch_display() if doc.batch else 'N/A',
                    'uploaded_by': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                    'uploaded_at': doc.created_at.strftime('%Y-%m-%d %H:%M'),
                    'archived_at': doc.updated_at.strftime('%Y-%m-%d %H:%M'),
                })
        
        # Get archived Policy Documents
        if not document_type or document_type == 'policy':
            policy_docs = PolicyDocument.objects.select_related('category', 'program', 'uploaded_by').filter(status='archived')
            for doc in policy_docs:
                results.append({
                    'id': doc.id,
                    'document_type': 'policy',
                    'document_name': doc.title,
                    'description': doc.description or '',
                    'category': doc.category.name if doc.category else 'N/A',
                    'program': doc.program.program_name if doc.program else 'N/A',
                    'batch': doc.get_batch_display() if doc.batch else 'N/A',
                    'uploaded_by': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                    'uploaded_at': doc.created_at.strftime('%Y-%m-%d %H:%M'),
                    'archived_at': doc.updated_at.strftime('%Y-%m-%d %H:%M'),
                })
        
        # Get archived LMSTC Documents
        if not document_type or document_type == 'lmstc_document':
            lmstc_docs = LMSTC_Documents.objects.select_related('program', 'uploaded_by').filter(status='archived')
            for doc in lmstc_docs:
                results.append({
                    'id': doc.id,
                    'document_type': 'lmstc_document',
                    'document_name': doc.document_name,
                    'description': doc.description or '',
                    'category': doc.get_document_type_display(),
                    'program': doc.program.program_name if doc.program else 'N/A',
                    'batch': doc.get_batch_display() if doc.batch else 'N/A',
                    'uploaded_by': doc.uploaded_by.get_full_name() or doc.uploaded_by.username,
                    'uploaded_at': doc.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                    'archived_at': doc.updated_at.strftime('%Y-%m-%d %H:%M'),
                })
        
        # Sort by archived date (newest first)
        results.sort(key=lambda x: x['archived_at'], reverse=True)
        
        # Pagination
        total_count = len(results)
        start_index = (page - 1) * per_page
        end_index = start_index + per_page
        paginated_results = results[start_index:end_index]
        
        return JsonResponse({
            'success': True,
            'documents': paginated_results,
            'total': total_count,
            'page': page,
            'per_page': per_page,
            'total_pages': (total_count + per_page - 1) // per_page
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to retrieve archived documents: {str(e)}'
        })


@csrf_exempt
@require_POST
def import_2024_excel_to_documents(request):
    """Import data from 2024.xlsx file into LMSTC_Documents"""
    try:
        # Check if user is authenticated and is staff/admin
        if not request.user.is_authenticated:
            return JsonResponse({'success': False, 'error': 'Authentication required'})
        
        # Path to the 2024.xlsx file
        file_path = os.path.join(settings.BASE_DIR, 'Applicant', 'static', 'data', '2024.xlsx')
        
        # Also check in staticfiles
        if not os.path.exists(file_path):
            file_path = os.path.join(settings.BASE_DIR, 'staticfiles', 'data', '2024.xlsx')
        
        if not os.path.exists(file_path):
            return JsonResponse({'success': False, 'error': '2024.xlsx file not found in static/data or staticfiles/data'})
        
        # Read Excel file
        try:
            wb = load_workbook(filename=file_path, data_only=True)
            ws = wb.active
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Failed to read 2024.xlsx: {str(e)}'
            })
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return JsonResponse({'success': False, 'error': '2024.xlsx is empty'})
        
        # Get header row
        header = [str(h).strip() if h is not None else '' for h in rows[0]]
        
        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []
        
        # Set a default date in 2024 for all imported records
        default_upload_date = datetime(2024, 1, 1, 12, 0, 0)
        default_upload_date = timezone.make_aware(default_upload_date)
        
        # Process each row
        for row_idx, row in enumerate(rows[1:], start=2):  # Start at 2 because row 1 is header
            try:
                if row is None:
                    continue
                
                # Create record dictionary from row
                record = {}
                for idx, value in enumerate(row):
                    col_name = header[idx] if idx < len(header) else f'Column{idx+1}'
                    if col_name:
                        record[col_name] = '' if value is None else str(value).strip()
                
                # Skip completely empty rows
                if not any(str(v).strip() for v in record.values() if v):
                    skipped_count += 1
                    continue
                
                # Extract key fields for matching
                first_name = record.get('first_name', '') or record.get('First Name', '') or record.get('firstname', '') or ''
                last_name = record.get('last_name', '') or record.get('Last Name', '') or record.get('lastname', '') or ''
                email = record.get('email', '') or record.get('Email', '') or ''
                
                # Try to find matching Learner_Profile
                learner_profile = None
                applicant = None
                
                if email:
                    try:
                        learner_profile = Learner_Profile.objects.filter(email__iexact=email).first()
                        if learner_profile and learner_profile.user:
                            applicant = learner_profile.user
                    except Exception:
                        pass
                
                # If no match by email, try by name
                if not learner_profile and first_name and last_name:
                    try:
                        learner_profile = Learner_Profile.objects.filter(
                            first_name__iexact=first_name,
                            last_name__iexact=last_name
                        ).first()
                        if learner_profile and learner_profile.user:
                            applicant = learner_profile.user
                    except Exception:
                        pass
                
                # Create document name from available data
                if first_name and last_name:
                    document_name = f"{first_name} {last_name} - Profile (2024)"
                elif email:
                    document_name = f"Profile - {email} (2024)"
                else:
                    document_name = f"Imported Record {row_idx} (2024)"
                
                # Create description from record data
                description_parts = []
                for key, value in record.items():
                    if value and key.lower() not in ['first_name', 'last_name', 'email', 'first name', 'last name']:
                        description_parts.append(f"{key}: {value}")
                description = "; ".join(description_parts[:10])  # Limit to first 10 fields
                
                # Check if document already exists (to avoid duplicates)
                existing_doc = None
                if learner_profile:
                    existing_doc = LMSTC_Documents.objects.filter(
                        learner_profile=learner_profile,
                        document_name__icontains="2024",
                        uploaded_at__year=2024
                    ).first()
                elif applicant:
                    existing_doc = LMSTC_Documents.objects.filter(
                        applicant=applicant,
                        document_name__icontains="2024",
                        uploaded_at__year=2024
                    ).first()
                
                if existing_doc:
                    # Update existing docwument
                    existing_doc.description = description
                    existing_doc.uploaded_at = default_upload_date
                    existing_doc.save()
                    updated_count += 1
                else:
                    # Create new document record
                    # Note: We're creating a document record without an actual file
                    # The document_file field will be empty, but the record will exist in the database
                    document = LMSTC_Documents.objects.create(
                        document_name=document_name,
                        document_type='applicant_profile',
                        description=description,
                        learner_profile=learner_profile,
                        applicant=applicant,
                        uploaded_by=request.user,
                        uploaded_at=default_upload_date,
                        status='active',
                        file_size=0,
                        mime_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    created_count += 1
                    
            except Exception as row_error:
                errors.append(f"Row {row_idx}: {str(row_error)}")
                skipped_count += 1
                continue
        
        # Prepare response
        response_data = {
            'success': True,
            'message': f'Import completed. {created_count} records created, {updated_count} records updated, {skipped_count} rows skipped.',
            'created_count': created_count,
            'updated_count': updated_count,
            'skipped_count': skipped_count,
            'total_processed': len(rows) - 1  # Exclude header
        }
        
        if errors:
            response_data['errors'] = errors[:10]  # Limit to first 10 errors
            response_data['error_count'] = len(errors)
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Import failed: {str(e)}'
        })
