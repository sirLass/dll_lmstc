from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .models import ApprovedApplicant, Learner_Profile, Programs, BatchCycle
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
from django.conf import settings


@login_required
@require_http_methods(["POST"])
def download_bulk_documents_excel(request):
    """
    Generate Excel file with selected document data using the 2024.xlsx template format.
    Accepts a list of selected document IDs (both LMSTC-* and LP-*) and filters.
    
    Fixes applied:
    - QuerySet properly converted to list to avoid evaluation issues
    - Documents are no longer skipped when filters are empty or when approval records are missing
    - Better learner_profile lookup - tries multiple methods to find associated profile
    - All selected documents are included in export, even if some data is missing
    """
    import json
    from .models import LMSTC_Documents, Learner_Profile, ApprovedApplicant, BatchCycle
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        # Get data from request
        data = json.loads(request.body)
        selected_ids = data.get('selected_ids', [])
        filters = data.get('filters', {})
        
        logger.info(f'Bulk download request - Selected IDs: {selected_ids}')
        logger.info(f'Filters: {filters}')
        
        if not selected_ids:
            return JsonResponse({'error': 'No documents selected'}, status=400)
        
        # Parse IDs to separate LMSTC documents and Learner Profiles
        lmstc_ids = []
        learner_profile_ids = []
        
        for id_val in selected_ids:
            try:
                id_str = str(id_val).strip()
                
                # Check if it's an LMSTC document ID (format: LMSTC-123 or just 123)
                if id_str.startswith('LMSTC-'):
                    numeric_id = int(id_str.replace('LMSTC-', ''))
                    lmstc_ids.append(numeric_id)
                # Check if it's a Learner Profile ID (format: LP-123)
                elif id_str.startswith('LP-'):
                    numeric_id = int(id_str.replace('LP-', ''))
                    learner_profile_ids.append(numeric_id)
                # If no prefix, assume it's an LMSTC document ID for backward compatibility
                else:
                    numeric_id = int(id_str)
                    lmstc_ids.append(numeric_id)
                    
            except (ValueError, AttributeError) as e:
                logger.warning(f'Could not parse ID {id_val}: {e}')
                continue
        
        logger.info(f'Parsed - LMSTC IDs: {lmstc_ids}, Learner Profile IDs: {learner_profile_ids}')
        
        if not lmstc_ids and not learner_profile_ids:
            return JsonResponse({'error': 'No valid document IDs provided'}, status=400)
        
        # Load the template file
        template_path = os.path.join(settings.BASE_DIR, 'Applicant', 'static', 'data', '2024.xlsx')
        
        try:
            template_wb = openpyxl.load_workbook(template_path)
            # Use the first sheet as template
            template_sheet = template_wb.worksheets[0]
        except Exception as e:
            logger.error(f'Could not load template: {e}')
            return JsonResponse({'error': f'Could not load template: {str(e)}'}, status=500)
        
        # Create new workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Get current year
        current_year = datetime.now().year
        
        # Create 3 trimester sheets like the template
        trimester_sheets = {}
        for trimester_num in [1, 2, 3]:
            sheet_name = f'{trimester_num}st Trimester {current_year}' if trimester_num == 1 else \
                        f'{trimester_num}nd Trimester {current_year}' if trimester_num == 2 else \
                        f'{trimester_num}rd Trimester {current_year}'
            trimester_sheets[trimester_num] = wb.create_sheet(title=sheet_name)
        
        # Copy template structure to all trimester sheets
        for trimester_num, ws in trimester_sheets.items():
            # Copy the template header structure (rows 1-7)
            for row_idx in range(1, 8):
                for col_idx in range(1, template_sheet.max_column + 1):
                    source_cell = template_sheet.cell(row_idx, col_idx)
                    target_cell = ws.cell(row_idx, col_idx)
                    
                    # Copy value
                    if source_cell.value:
                        target_cell.value = source_cell.value
                    
                    # Copy formatting
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy()
                        target_cell.border = source_cell.border.copy()
                        target_cell.fill = source_cell.fill.copy()
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = source_cell.protection.copy()
                        target_cell.alignment = source_cell.alignment.copy()
            
            # Copy merged cells from template header
            for merged_cell_range in template_sheet.merged_cells.ranges:
                if merged_cell_range.min_row <= 7:
                    ws.merge_cells(str(merged_cell_range))
            
            # Copy column dimensions
            for col_letter in template_sheet.column_dimensions:
                if col_letter in ws.column_dimensions:
                    ws.column_dimensions[col_letter].width = template_sheet.column_dimensions[col_letter].width
            
            # Update title in Row 2 for each trimester
            trimester_label = '1ST' if trimester_num == 1 else '2ND' if trimester_num == 2 else '3RD'
            ws['A2'] = f'{trimester_label} TRIMESTER {current_year}'
            
            # Override header labels to start with applicant personal info, then region/provider info, then program info
            # Note: Template rows 1-7 are header structure; we update row 5 labels commonly used for column headers
            try:
                ws.cell(5, 1, 'Family/ Last Name')
                ws.cell(5, 2, 'First Name')
                ws.cell(5, 3, 'Middle Name')
                ws.cell(5, 4, 'ULI')
                ws.cell(5, 5, 'Contact Number')
                ws.cell(5, 6, 'E-mail Address')
                ws.cell(5, 7, 'Region')
                ws.cell(5, 8, 'Province')
                ws.cell(5, 9, 'Congressional District')
                ws.cell(5, 10, 'Municipality/ City')
                ws.cell(5, 11, 'Name of Provider')
                ws.cell(5, 12, 'Complete Address of Provider/Training Venue')
                ws.cell(5, 13, 'Type of Provider')
                ws.cell(5, 14, 'Classification of Provider')
                ws.cell(5, 15, 'Industry Sector of Qualification')
                ws.cell(5, 16, 'TVET Program Registration Status')
                ws.cell(5, 17, 'Qualification/ Program Title')
                ws.cell(5, 18, 'CoPR Number')
                ws.cell(5, 19, 'Training Calendar Code')
                ws.cell(5, 20, 'Delivery Mode')
            except Exception:
                # If header row indices differ in template, ignore label override rather than failing
                pass
        
        # Fetch selected LMSTC documents - ensure we get a list
        lmstc_docs = []
        if lmstc_ids:
            lmstc_docs = list(LMSTC_Documents.objects.select_related(
                'learner_profile', 'applicant', 'uploaded_by', 'learner_profile__user'
            ).filter(id__in=lmstc_ids, status='active'))
        
        # Fetch selected Learner Profiles - ensure we get a list
        learner_profiles = []
        if learner_profile_ids:
            learner_profiles = list(Learner_Profile.objects.select_related('user').filter(
            id__in=learner_profile_ids
            ))
            logger.info(f'Queried for learner profile IDs: {learner_profile_ids}')
            logger.info(f'Found learner profiles: {[lp.id for lp in learner_profiles]}')
        
        logger.info(f'Found {len(lmstc_docs)} LMSTC documents and {len(learner_profiles)} learner profiles')
        
        if len(learner_profiles) == 0 and learner_profile_ids:
            logger.warning(f'No learner profiles found for IDs: {learner_profile_ids}')
        if len(lmstc_docs) == 0 and lmstc_ids:
            logger.warning(f'No LMSTC documents found for IDs: {lmstc_ids}')
        
        # Apply filters from the filter parameters
        filter_batch = filters.get('batch', '')
        filter_year = filters.get('year', '')
        filter_program = filters.get('program', '')
        filter_search = filters.get('search', '')
        
        # Create a combined list of records to process
        # Each record will have: learner_profile, approved_applicant, trimester
        records_by_trimester = {1: [], 2: [], 3: []}
        
        # Process LMSTC Documents
        for doc in lmstc_docs:
            learner = doc.learner_profile
            
            # Try to get learner_profile if not directly linked
            if not learner and doc.applicant:
                try:
                    learner = Learner_Profile.objects.filter(user=doc.applicant).first()
                except Exception:
                    pass
            
            trimester = 1  # Default to 1st trimester
            approved_app = None
            
            # Try to get ApprovedApplicant for BatchCycle info
            applicant_user = doc.applicant
            if not applicant_user and learner and getattr(learner, 'user', None):
                applicant_user = learner.user
            
            if applicant_user:
                try:
                    approved_app = ApprovedApplicant.objects.select_related('program').filter(
                        applicant=applicant_user
                    ).order_by('-approved_at').first()
                    if not approved_app:
                        # Fallback to walk-in approval
                        from .models import ApprovedWalkIn
                        walkin_app = ApprovedWalkIn.objects.filter(applicant=applicant_user).order_by('-approved_at').first()
                        if walkin_app:
                            # Apply filters only if they are explicitly set (not empty strings)
                            if filter_batch and str(walkin_app.batch_number) != str(filter_batch):
                                logger.info(f'Skipping doc {doc.id} - walk-in batch mismatch: {walkin_app.batch_number} != {filter_batch}')
                                continue
                            if filter_year and walkin_app.enrollment_year and str(walkin_app.enrollment_year) != str(filter_year):
                                logger.info(f'Skipping doc {doc.id} - walk-in year mismatch: {walkin_app.enrollment_year} != {filter_year}')
                                continue
                            # Derive trimester from nothing (default 1)
                            approved_app = None
                    else:
                        # Apply BatchCycle filters only if they are explicitly set
                        if filter_batch and str(approved_app.batch_number) != str(filter_batch):
                            logger.info(f'Skipping doc {doc.id} - batch mismatch: {approved_app.batch_number} != {filter_batch}')
                            continue
                        if filter_year and approved_app.enrollment_year and str(approved_app.enrollment_year) != str(filter_year):
                            logger.info(f'Skipping doc {doc.id} - year mismatch: {approved_app.enrollment_year} != {filter_year}')
                            continue
                        if filter_program and approved_app.program_id and str(approved_app.program_id) != str(filter_program):
                            logger.info(f'Skipping doc {doc.id} - program mismatch: {approved_app.program_id} != {filter_program}')
                            continue
                        # Determine trimester from enrollment_semester
                        if approved_app.enrollment_semester:
                            try:
                                trimester = int(approved_app.enrollment_semester)
                                if trimester not in [1, 2, 3]:
                                    trimester = 1
                            except (ValueError, TypeError):
                                trimester = 1
                except Exception as e:
                    logger.warning(f'Error getting approval record for doc {doc.id}: {e}')
                    # Don't skip - continue with default values
            
            # If no applicant, try to use batch from doc for trimester
            if not applicant_user or not approved_app:
                if doc.batch:
                    try:
                        batch_str = str(doc.batch)
                        if batch_str.startswith('batch_'):
                            batch_num = int(batch_str.replace('batch_', ''))
                        else:
                            batch_num = int(batch_str)
                        if batch_num in [1, 2, 3]:
                            trimester = batch_num
                    except (ValueError, IndexError):
                        pass
            
            # Always add the record - don't skip documents just because they don't have approval records
            # The user selected these documents, so they should be included
            records_by_trimester[trimester].append({
                'learner_profile': learner,
                'approved_applicant': approved_app,
                'document': doc
            })
        
        # Process standalone Learner Profiles
        for learner in learner_profiles:
            trimester = 1  # Default
            approved_app = None
            
            # Try to find ApprovedApplicant for this learner
            if learner.user:
                try:
                    approved_app = ApprovedApplicant.objects.select_related('program').filter(
                        applicant=learner.user
                    ).order_by('-approved_at').first()
                    if not approved_app:
                        from .models import ApprovedWalkIn
                        walkin_app = ApprovedWalkIn.objects.filter(applicant=learner.user).order_by('-approved_at').first()
                        if walkin_app:
                            # Apply filters only if explicitly set
                            if filter_batch and str(walkin_app.batch_number) != str(filter_batch):
                                logger.info(f'Skipping learner {learner.id} - walk-in batch mismatch')
                                continue
                            if filter_year and walkin_app.enrollment_year and str(walkin_app.enrollment_year) != str(filter_year):
                                logger.info(f'Skipping learner {learner.id} - walk-in year mismatch')
                                continue
                            # trimester remains default
                    else:
                        # Apply BatchCycle filters only if explicitly set
                        if filter_batch and str(approved_app.batch_number) != str(filter_batch):
                            logger.info(f'Skipping learner {learner.id} - batch mismatch')
                            continue
                        if filter_year and approved_app.enrollment_year and str(approved_app.enrollment_year) != str(filter_year):
                            logger.info(f'Skipping learner {learner.id} - year mismatch')
                            continue
                        if filter_program and approved_app.program_id and str(approved_app.program_id) != str(filter_program):
                            logger.info(f'Skipping learner {learner.id} - program mismatch')
                            continue
                        
                        # Determine trimester
                        if approved_app.enrollment_semester:
                            try:
                                trimester = int(approved_app.enrollment_semester)
                                if trimester not in [1, 2, 3]:
                                    trimester = 1
                            except (ValueError, TypeError):
                                trimester = 1
                            
                except Exception as e:
                    logger.warning(f'Error getting approval record for learner {learner.id}: {e}')
                    # Don't skip - continue with default values
            
            # Always add the record - user selected these profiles, so they should be included
            records_by_trimester[trimester].append({
                'learner_profile': learner,
                'approved_applicant': approved_app,
                'document': None
            })
        
        logger.info(f'Records by trimester after filtering: T1={len(records_by_trimester[1])}, T2={len(records_by_trimester[2])}, T3={len(records_by_trimester[3])}')
        
        # Process each trimester sheet
        for trimester_num, ws in trimester_sheets.items():
            records_for_trimester = records_by_trimester[trimester_num]
            data_row = 8  # Start writing data from row 8
            
            for record in records_for_trimester:
                try:
                    # Extract record components
                    learner = record['learner_profile']
                    approved_app = record['approved_applicant']
                    doc = record.get('document')
                    
                    # If no learner but we have a document with applicant, try to get learner_profile one more time
                    if not learner and doc and doc.applicant:
                        try:
                            learner = Learner_Profile.objects.filter(user=doc.applicant).first()
                        except Exception:
                            pass
                    
                    # Use Learner_Profile data (should always be available now)
                    if learner:
                        # Trainee Personal Info (now columns a-f)
                        ws.cell(data_row, 1, (learner.last_name or 'N/A'))  # Family/ Last Name
                        ws.cell(data_row, 2, (learner.first_name or 'N/A'))  # First Name
                        ws.cell(data_row, 3, (learner.middle_name or ''))  # Middle Name
                        ws.cell(data_row, 4, 'N/A')  # ULI
                        ws.cell(data_row, 5, (learner.contact_number or 'N/A'))  # Contact Number
                        ws.cell(data_row, 6, (learner.email or 'N/A'))  # E-mail Address

                        # TVET Provider / Location Info (now columns g-n)
                        ws.cell(data_row, 7, learner.region_name or learner.region or 'Region IV-A')  # Region
                        ws.cell(data_row, 8, learner.province_name or learner.province or 'Quezon')  # Province
                        ws.cell(data_row, 9, learner.district or '2nd District')  # Congressional District
                        ws.cell(data_row, 10, learner.city_name or learner.city or 'Lucena City')  # Municipality/City
                        ws.cell(data_row, 11, 'Dalubhasaan ng Lungsod ng Lucena')  # Name of Provider
                        ws.cell(data_row, 12, 'Maharlika Hi-way, Brgy. Isabang Lucena City')  # Complete Address
                        ws.cell(data_row, 13, 'Public')  # Type of Provider
                        ws.cell(data_row, 14, 'N/A')  # Classification of Provider

                        # Program Info (now columns o-t)
                        ws.cell(data_row, 15, 'N/A')  # Industry Sector
                        ws.cell(data_row, 16, 'N/A')  # Program Status
                        if approved_app and approved_app.program:
                            ws.cell(data_row, 17, approved_app.program.program_name)  # Program
                        else:
                            ws.cell(data_row, 17, learner.course_or_qualification or 'N/A')  # Program
                        ws.cell(data_row, 18, 'N/A')  # CoPR
                        ws.cell(data_row, 19, 'N/A')  # Training Calendar
                        ws.cell(data_row, 20, 'Face-to-face')  # Delivery Mode
                        
                        # Address breakdown (columns t-x)
                        ws.cell(data_row, 21, learner.street or 'N/A')  # Street
                        ws.cell(data_row, 22, learner.barangay_name or learner.barangay or 'N/A')  # Barangay
                        ws.cell(data_row, 23, learner.city_name or learner.city or 'N/A')  # City
                        ws.cell(data_row, 24, learner.district or 'N/A')  # District
                        ws.cell(data_row, 25, learner.province_name or learner.province or 'N/A')  # Province
                        
                        # Demographics (columns y-ac and beyond)
                        ws.cell(data_row, 26, learner.sex or 'N/A')  # Sex (y)
                        ws.cell(data_row, 27, learner.birthdate if learner.birthdate else 'N/A')  # DOB (z)
                        ws.cell(data_row, 28, learner.age or 'N/A')  # Age (aa)
                        ws.cell(data_row, 29, learner.civil_status or 'N/A')  # Civil Status (ab)
                        ws.cell(data_row, 30, learner.educational_attainment or 'N/A')  # Education (ac)
                        
                        # Parent/Guardian (ad)
                        ws.cell(data_row, 31, learner.parent_guardian or 'N/A')
                        
                        # Client Classification (ae)
                        classifications = learner.classifications.all()
                        if classifications:
                            class_names = ', '.join([c.name for c in classifications])
                            ws.cell(data_row, 32, class_names)
                        else:
                            ws.cell(data_row, 32, learner.other_classification or 'N/A')
                        
                        # Employment Status (af)
                        ws.cell(data_row, 33, learner.employment_status or 'N/A')
                        
                        # Employment Type (ag) - Not in model, use N/A
                        ws.cell(data_row, 34, 'N/A')
                        
                        # Birthplace (ah-aj)
                        ws.cell(data_row, 35, learner.birthplace_regionb_name or 'N/A')  # Region
                        ws.cell(data_row, 36, learner.birthplace_provinceb_name or 'N/A')  # Province
                        ws.cell(data_row, 37, learner.birthplace_cityb_name or 'N/A')  # City
                        
                        # Disability (ak-al)
                        disability_types = learner.disability_types.all()
                        if disability_types:
                            ws.cell(data_row, 38, ', '.join([dt.name for dt in disability_types]))
                        else:
                            ws.cell(data_row, 38, 'None')
                        
                        disability_causes = learner.disability_causes.all()
                        if disability_causes:
                            ws.cell(data_row, 39, ', '.join([dc.name for dc in disability_causes]))
                        else:
                            ws.cell(data_row, 39, 'N/A')
                        
                        # Taken NCAE? (am) - Not in model
                        ws.cell(data_row, 40, 'N/A')
                        
                        # NCAE Rating (an-ao) - Not in model
                        ws.cell(data_row, 41, 'N/A')  # Rating
                        ws.cell(data_row, 42, 'N/A')  # Year Taken
                        
                        # Training/Course Info (ap-aq)
                        ws.cell(data_row, 43, learner.course_or_qualification or 'N/A')  # Course
                        ws.cell(data_row, 44, learner.scholarship_package or 'N/A')  # Scholarship
                        
                        # Entry Date (ar)
                        ws.cell(data_row, 45, learner.entry_date if learner.entry_date else 'N/A')
                        
                        # Exit Date (as) - Use date_accomplished
                        ws.cell(data_row, 46, learner.date_accomplished if learner.date_accomplished else 'N/A')
                        
                        # Assessment/Certification (at-aw) - Not in model
                        ws.cell(data_row, 47, 'N/A')  # Competency Assessment
                        ws.cell(data_row, 48, 'N/A')  # Certificate Number
                        ws.cell(data_row, 49, 'N/A')  # Date Issued
                        ws.cell(data_row, 50, 'N/A')  # Assessment Center
                        
                        # Employment After Training (ax-ba) - Use employment fields
                        ws.cell(data_row, 51, learner.employment_status or 'N/A')  # Employed after training
                        ws.cell(data_row, 52, learner.company_name or 'N/A')  # Company Name
                        ws.cell(data_row, 53, learner.monthly_income or 'N/A')  # Monthly Income
                        ws.cell(data_row, 54, learner.date_hired if learner.date_hired else 'N/A')  # Date Hired
                    else:
                        # NO Learner_Profile - This is a bulk uploaded document with metadata only
                        # Trainee Personal Info first
                        doc_name_parts = doc.document_name.replace('.pdf', '').replace('.xlsx', '').replace('.docx', '').split()
                        if doc.applicant:
                            ws.cell(data_row, 1, doc.applicant.last_name or 'N/A')
                            ws.cell(data_row, 2, doc.applicant.first_name or 'N/A')
                            ws.cell(data_row, 6, doc.applicant.email or 'N/A')
                        elif len(doc_name_parts) >= 2:
                            # Try to extract name from filename
                            ws.cell(data_row, 1, doc_name_parts[0])  # Assume first word is last name
                            ws.cell(data_row, 2, doc_name_parts[1])  # Assume second word is first name
                        else:
                            ws.cell(data_row, 1, f'Document {doc.id}')
                            ws.cell(data_row, 2, 'Bulk Upload')

                        ws.cell(data_row, 3, '')  # Middle Name
                        ws.cell(data_row, 4, 'N/A')  # ULI
                        ws.cell(data_row, 5, 'N/A')  # Contact

                        # Provider / Location Info next
                        ws.cell(data_row, 7, 'Region IV-A')
                        ws.cell(data_row, 8, 'Quezon')
                        ws.cell(data_row, 9, '2nd District')
                        ws.cell(data_row, 10, 'Lucena City')
                        ws.cell(data_row, 11, 'Dalubhasaan ng Lungsod ng Lucena')
                        ws.cell(data_row, 12, 'Maharlika Hi-way, Brgy. Isabang Lucena City')
                        ws.cell(data_row, 13, 'Public')
                        ws.cell(data_row, 14, 'N/A')

                        # Program Info
                        ws.cell(data_row, 15, 'N/A')  # Industry Sector
                        ws.cell(data_row, 16, 'N/A')  # Program Status
                        if approved_app and approved_app.program:
                            ws.cell(data_row, 17, approved_app.program.program_name)
                        elif doc.program:
                            ws.cell(data_row, 17, doc.program.program_name)
                        else:
                            ws.cell(data_row, 17, doc.get_document_type_display())
                        ws.cell(data_row, 18, 'N/A')  # CoPR
                        ws.cell(data_row, 19, 'N/A')  # Training Calendar
                        ws.cell(data_row, 20, 'Face-to-face')  # Delivery Mode
                        
                        # Log that this record has no learner data (shouldn't happen with new logic)
                        if doc:
                            logger.warning(f'Document {doc.id} has no learner_profile - using metadata only')
                        else:
                            logger.warning(f'Record has no learner_profile and no document')
                    
                    data_row += 1
                    
                except Exception as e:
                    logger.error(f'Error processing record: {e}', exc_info=True)
                    continue
            
            # Log results for this trimester
            records_added = data_row - 8
            logger.info(f'Trimester {trimester_num}: {records_added} records added')
            
            if records_added == 0:
                ws.cell(8, 1, 'No data for this trimester')
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'DLL_LMSTC_Documents_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        
        return response
    
    except json.JSONDecodeError as e:
        logger.error(f'JSON decode error: {e}')
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f'Error generating Excel: {e}', exc_info=True)
        return JsonResponse({'error': f'Failed to generate Excel: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def download_applicants_excel(request):
    """
    Generate Excel file with applicant data based on filtered criteria.
    Uses the format from 2024.xlsx template with tabs for each trimester.
    """
    import json
    
    # Get filters from request
    data = json.loads(request.body)
    filters = data.get('filters', {})
    
    # Load the template file
    template_path = os.path.join(settings.BASE_DIR, 'Applicant', 'static', 'data', '2024.xlsx')
    
    try:
        template_wb = openpyxl.load_workbook(template_path)
    except Exception as e:
        return JsonResponse({'error': f'Could not load template: {str(e)}'}, status=500)
    
    # Create new workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Get current year for the sheets
    current_year = datetime.now().year
    
    # Define trimester/semester mapping
    trimesters = [
        ('1', f'1st Trimester {current_year}'),
        ('2', f'2nd Trimester {current_year}'),
        ('3', f'3rd Trimester {current_year}')
    ]
    
    # Query base approved applicants
    approved_applicants = ApprovedApplicant.objects.select_related(
        'applicant', 'program'
    ).filter(status='active')
    
    # Apply filters
    batch_filter = filters.get('batch', '')
    year_filter = filters.get('year', '')
    program_filter = filters.get('program', '')
    search_filter = filters.get('search', '')
    
    if batch_filter:
        approved_applicants = approved_applicants.filter(batch_number=batch_filter)
    
    if year_filter:
        approved_applicants = approved_applicants.filter(enrollment_year=int(year_filter))
    
    if program_filter:
        approved_applicants = approved_applicants.filter(program_id=int(program_filter))
    
    if search_filter:
        approved_applicants = approved_applicants.filter(
            Q(applicant__username__icontains=search_filter) |
            Q(applicant__first_name__icontains=search_filter) |
            Q(applicant__last_name__icontains=search_filter)
        )
    
    # Process each trimester
    for semester_num, sheet_name in trimesters:
        # Filter applicants for this semester
        semester_applicants = approved_applicants.filter(enrollment_semester=semester_num)
        
        # Create sheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Copy headers from template
        template_sheet_name = f'{semester_num}st Trimester 2024' if semester_num == '1' else (
            f'{semester_num}nd Trimester 2024' if semester_num == '2' else f'{semester_num}rd Trimester 2024'
        )
        
        if template_sheet_name in template_wb.sheetnames:
            template_ws = template_wb[template_sheet_name]
            
            # Copy first 7 rows (headers, titles, etc.)
            for row_num in range(1, 8):
                for col_num in range(1, template_ws.max_column + 1):
                    source_cell = template_ws.cell(row_num, col_num)
                    target_cell = ws.cell(row_num, col_num)
                    
                    # Copy value
                    target_cell.value = source_cell.value
                    
                    # Copy styles
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy()
                        target_cell.border = source_cell.border.copy()
                        target_cell.fill = source_cell.fill.copy()
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = source_cell.protection.copy()
                        target_cell.alignment = source_cell.alignment.copy()
                    
                    # Copy column width
                    if row_num == 1:
                        column_letter = openpyxl.utils.get_column_letter(col_num)
                        ws.column_dimensions[column_letter].width = template_ws.column_dimensions[column_letter].width
                    
                    # Copy merged cells
                    if row_num == 1:
                        for merged_range in template_ws.merged_cells.ranges:
                            ws.merge_cells(str(merged_range))
        else:
            # Create basic header structure if template not found
            ws['A1'] = 'DLL-LMSTC'
            ws['A2'] = f'{sheet_name.upper()}'
            ws['A3'] = f'TRAINING DURATION - {datetime.now().strftime("%B %d, %Y")}'
            
            # Add header row
            headers = [
                'Region', 'Province', 'Congressional District', 'Municipality/ City',
                'Name of Provider', 'Complete Address of Provider/Training Venue',
                'Type of Provider', 'Classification of Provider',
                'Industry Sector of Qualification', 'TVET Program Registration Status',
                'Qualification/ Program Title', 'CoPR Number', 'Training Calendar Code',
                'Delivery Mode', 'Family/ Last Name', 'First Name', 'Middle Name',
                'ULI', 'Contact Number', 'E-mail Address', 'Street Address',
                'Barangay', 'Municipality/ City', 'District', 'Province', 'Sex',
                'Date of Birth', 'Age', 'Civil Status', 'Highest Educational Attainment'
            ]
            
            for col_num, header in enumerate(headers, start=1):
                ws.cell(5, col_num, header)
        
        # Add data starting from row 8
        data_row = 8
        
        for approved in semester_applicants:
            try:
                learner = Learner_Profile.objects.filter(user=approved.applicant).first()
                
                if learner:
                    # Column mapping based on template
                    row_data = {
                        1: 'Region IV- A',  # Region (a)
                        2: learner.province_name or 'Quezon',  # Province (b)
                        3: '2nd District',  # Congressional District (c)
                        4: learner.city_name or 'Lucena City',  # Municipality/City (d)
                        5: 'Dalubhasaan ng Lungsod ng Lucena',  # Name of Provider (e)
                        6: 'Maharlika Hi-way, Brgy. Isabang Lucena City',  # Complete Address (f)
                        7: 'Public',  # Type of Provider (g)
                        8: 'N/A',  # Classification of Provider (h)
                        9: 'N/A',  # Industry Sector (i)
                        10: 'N/A',  # TVET Program Registration Status (j)
                        11: approved.program.program_name if approved.program else 'N/A',  # Qualification/Program Title (k)
                        12: 'N/A',  # CoPR Number (l)
                        13: 'N/A',  # Training Calendar Code (m)
                        14: 'Face-to-face',  # Delivery Mode (n)
                        15: learner.last_name.upper(),  # Family/Last Name (o)
                        16: learner.first_name.upper(),  # First Name (p)
                        17: learner.middle_name.upper() if learner.middle_name else '',  # Middle Name (q)
                        18: 'N/A',  # ULI (r)
                        19: learner.contact_number or 'N/A',  # Contact Number (s)
                        20: learner.email or 'N/A',  # E-mail Address (t)
                        21: learner.street or '',  # Street Address (u)
                        22: learner.barangay_name or '',  # Barangay (v)
                        23: learner.city_name or '',  # Municipality/City (w)
                        24: learner.district or '',  # District (x)
                        25: learner.province_name or '',  # Province (y)
                        26: learner.sex or 'N/A',  # Sex (z)
                        27: learner.birthdate if learner.birthdate else '',  # Date of Birth (aa)
                        28: learner.age if learner.age else '',  # Age (ab)
                        29: learner.civil_status or 'N/A',  # Civil Status (ac)
                        30: learner.educational_attainment or 'N/A',  # Highest Educational Attainment (ad)
                    }
                    
                    # Write data to cells
                    for col_num, value in row_data.items():
                        ws.cell(data_row, col_num, value)
                    
                    data_row += 1
            
            except Exception as e:
                # Skip problematic records but continue processing
                print(f"Error processing applicant {approved.id}: {str(e)}")
                continue
    
    # Remove empty sheets (trimesters with no data)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row <= 7:  # Only headers, no data
            # Keep the sheet but add a note
            ws.cell(8, 1, 'No applicants enrolled for this trimester')
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'Applicants_Report_{current_year}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response
