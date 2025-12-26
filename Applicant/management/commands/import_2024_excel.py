from django.core.management.base import BaseCommand
from django.conf import settings
from django.utils import timezone
from datetime import datetime, date
import os
from openpyxl import load_workbook

from Applicant.models import LMSTC_Documents, Learner_Profile, Applicant, Programs


class Command(BaseCommand):
    help = 'Import data from 2024.xlsx file into LMSTC_Documents and Learner_Profile'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            help='Path to the Excel file (default: Applicant/static/data/2024.xlsx)',
        )
        parser.add_argument(
            '--user-id',
            type=int,
            help='User ID to use as uploaded_by (default: first admin user)',
        )

    def get_batch_from_sheet_name(self, sheet_name):
        """Map sheet name to batch number"""
        sheet_lower = sheet_name.lower().strip()
        if '1st' in sheet_lower or 'first' in sheet_lower:
            return 'batch_1'
        elif '2nd' in sheet_lower or 'second' in sheet_lower:
            return 'batch_2'
        elif '3rd' in sheet_lower or 'third' in sheet_lower:
            return 'batch_3'
        else:
            # Default to batch_1 if can't determine
            return 'batch_1'

    def extract_name_parts(self, record):
        """Extract first name, last name, and middle name from record"""
        # Try various column name variations (case-insensitive search)
        def get_value(keys):
            for key in keys:
                # Try exact match first
                if key in record:
                    val = record[key]
                    if val and str(val).strip():
                        return str(val).strip()
                # Try case-insensitive match
                for k, v in record.items():
                    if k.lower() == key.lower():
                        if v and str(v).strip():
                            return str(v).strip()
            return ''
        
        last_name = get_value([
            'last_name', 'Last Name', 'lastname', 'Last', 'Family/ Last Name',
            'Family Name', 'Family', 'Surname', 'surname', 'Last Name / Surname'
        ])
        
        first_name = get_value([
            'first_name', 'First Name', 'firstname', 'First', 'Given Name',
            'Given', 'First Name / Given Name'
        ])
        
        middle_name = get_value([
            'middle_name', 'Middle Name', 'middlename', 'Middle', 'Middle Initial',
            'M.I.', 'MI'
        ])
        
        # If full name is in one field, try to split it
        if not first_name and not last_name:
            full_name = get_value([
                'full_name', 'Full Name', 'name', 'Name', 'Trainee Name',
                'Learner Name', 'Complete Name'
            ])
            if full_name:
                parts = full_name.split()
                if len(parts) >= 2:
                    last_name = parts[0]
                    first_name = parts[1]
                    if len(parts) > 2:
                        middle_name = ' '.join(parts[2:])
        
        return first_name, last_name, middle_name

    def get_full_name_display(self, first_name, last_name, middle_name):
        """Format full name as 'Last Name, First Name Middle Name'"""
        parts = []
        if last_name:
            parts.append(last_name)
        if first_name:
            parts.append(first_name)
        if middle_name:
            parts.append(middle_name)
        return ', '.join(parts) if parts else 'Unknown'

    def find_or_create_program(self, program_name):
        """Find existing program using flexible matching"""
        if not program_name or not program_name.strip():
            return None
        
        program_name_clean = program_name.strip()
        
        # Normalize the search term (remove hyphens, convert to lowercase)
        normalized_search = program_name_clean.lower().replace('-', ' ').replace('_', ' ')
        
        # Common abbreviation mappings
        abbreviation_map = {
            'smaw': 'shielded metal arc welding',
            'bp': 'bread and pastry',
            'eim': 'electrical installation and maintenance',
            'epas': 'electronic products assembly and service',
            'rac': 'rac servicing',
            'domrac': 'rac servicing',
            'aea': 'auto electrical assembly',
            'ama': 'auto mechanical assembly',
        }
        
        # Check if it's an abbreviation
        search_lower = normalized_search
        for abbrev, full_name in abbreviation_map.items():
            if abbrev in search_lower:
                search_lower = full_name
                break
        
        # Try exact match first
        try:
            program = Programs.objects.filter(program_name__iexact=program_name_clean).first()
            if program:
                return program
        except Exception:
            pass
        
        # Try contains match (normalized)
        try:
            all_programs = Programs.objects.all()
            for prog in all_programs:
                prog_normalized = prog.program_name.lower().replace('-', ' ').replace('_', ' ')
                # Remove common suffixes for matching
                prog_clean = prog_normalized.replace(' nc ii', '').replace(' ncii', '').replace(' nc2', '').strip()
                search_clean = search_lower.replace(' nc ii', '').replace(' ncii', '').replace(' nc2', '').strip()
                
                # Check if search term is in program name
                if search_clean in prog_clean or prog_clean in search_clean:
                    return prog
                # Check if program name is in search term
                if prog_clean in search_clean or search_clean in prog_clean:
                    return prog
        except Exception:
            pass
        
        # Try word-based matching (at least 2 significant words match)
        try:
            search_words = [w for w in search_lower.split() if len(w) > 2]  # Ignore short words
            if len(search_words) >= 2:
                all_programs = Programs.objects.all()
                best_match = None
                best_score = 0
                
                for prog in all_programs:
                    prog_normalized = prog.program_name.lower().replace('-', ' ').replace('_', ' ')
                    prog_words = [w for w in prog_normalized.split() if len(w) > 2]
                    
                    # Count matching words
                    matching_words = sum(1 for word in search_words if any(word in pw or pw in word for pw in prog_words))
                    if matching_words >= 2 and matching_words > best_score:
                        best_score = matching_words
                        best_match = prog
                
                if best_match:
                    return best_match
        except Exception:
            pass
        
        return None

    def handle(self, *args, **options):
        # Get file path
        file_path = options.get('file')
        if not file_path:
            file_path = os.path.join(settings.BASE_DIR, 'Applicant', 'static', 'data', '2024.xlsx')
            # Also check in staticfiles
            if not os.path.exists(file_path):
                file_path = os.path.join(settings.BASE_DIR, 'staticfiles', 'data', '2024.xlsx')
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'File not found: {file_path}'))
            return
        
        # Get user for uploaded_by
        user_id = options.get('user_id')
        if user_id:
            try:
                user = Applicant.objects.get(id=user_id, is_staff=True)
            except Applicant.DoesNotExist:
                self.stdout.write(self.style.ERROR(f'User with ID {user_id} not found or is not staff'))
                return
        else:
            # Get first admin user
            user = Applicant.objects.filter(is_staff=True).first()
            if not user:
                self.stdout.write(self.style.ERROR('No admin user found. Please create an admin user first.'))
                return
        
        self.stdout.write(f'Using user: {user.username} (ID: {user.id})')
        self.stdout.write(f'Reading file: {file_path}')
        
        # Read Excel file
        try:
            wb = load_workbook(filename=file_path, data_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Failed to read Excel file: {str(e)}'))
            return
        
        # Process each sheet
        total_docs_created = 0
        total_docs_updated = 0
        total_profiles_created = 0
        total_profiles_updated = 0
        total_skipped = 0
        errors = []
        
        # Set a default date in 2024 for all imported records
        default_upload_date = datetime(2024, 1, 1, 12, 0, 0)
        default_upload_date = timezone.make_aware(default_upload_date)
        default_entry_date = date(2024, 1, 1)
        
        sheet_names = wb.sheetnames
        self.stdout.write(f'Found {len(sheet_names)} sheet(s): {", ".join(sheet_names)}')
        
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            batch = self.get_batch_from_sheet_name(sheet_name)
            self.stdout.write(f'\nProcessing sheet: "{sheet_name}" -> Batch: {batch}')
            
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                self.stdout.write(self.style.WARNING(f'  Sheet "{sheet_name}" is empty, skipping'))
                continue
            
            # Try to find the header row (look for common column names)
            # Prioritize rows that have "Family/ Last Name" or "Last Name" as these are data column headers
            header_row_idx = 0
            for idx, row in enumerate(rows[:10]):  # Check first 10 rows
                row_values = [str(v).strip() if v is not None else '' for v in row]
                row_values_lower = [v.lower() for v in row_values]
                # Check if this row contains actual data column headers (not section headers)
                # Look for specific column names that indicate data headers
                has_data_headers = any(
                    keyword in ' '.join(row_values_lower) 
                    for keyword in ['family/ last name', 'last name', 'first name', 'qualification/ program title']
                )
                # Also check for common header keywords but avoid section headers
                has_common_headers = any(
                    keyword in ' '.join(row_values_lower) 
                    for keyword in ['email', 'contact number', 'region', 'province']
                )
                if has_data_headers or (has_common_headers and 'profile' not in ' '.join(row_values_lower)):
                    header_row_idx = idx
                    break
            
            # Get header row
            header = [str(h).strip() if h is not None else '' for h in rows[header_row_idx]]
            data_start_row = header_row_idx + 1
            self.stdout.write(f'  Found header at row {header_row_idx + 1}')
            self.stdout.write(f'  Found {len(rows) - data_start_row} data rows')
            self.stdout.write(f'  Header columns: {", ".join([h for h in header[:15] if h])}...' if len(header) > 15 else f'  Header columns: {", ".join([h for h in header if h])}')
            
            # Process each row
            for row_idx, row in enumerate(rows[data_start_row:], start=data_start_row + 1):
                try:
                    if row is None:
                        continue
                    
                    # Create record dictionary from row
                    record = {}
                    for idx, value in enumerate(row):
                        col_name = header[idx] if idx < len(header) else f'Column{idx+1}'
                        if col_name:
                            record[col_name] = '' if value is None else str(value).strip()
                    
                    # Skip completely empty rows
                    if not any(str(v).strip() for v in record.values() if v):
                        total_skipped += 1
                        continue
                    
                    # Extract name parts
                    first_name, last_name, middle_name = self.extract_name_parts(record)
                    
                    # Skip if no name found
                    if not first_name and not last_name:
                        if row_idx <= 3:  # Debug first few rows
                            self.stdout.write(self.style.WARNING(f'    Row {row_idx}: No name found. Record keys: {list(record.keys())[:5]}'))
                        total_skipped += 1
                        continue
                    
                    # Extract other fields (case-insensitive)
                    def get_field_value(keys):
                        for key in keys:
                            if key in record:
                                val = record[key]
                                if val and str(val).strip():
                                    return str(val).strip()
                            for k, v in record.items():
                                if k.lower() == key.lower():
                                    if v and str(v).strip():
                                        return str(v).strip()
                        return ''
                    
                    email = get_field_value(['email', 'Email', 'E-mail', 'e-mail', 'Email Address', 'E-mail Address/ Facebook Account/ Twitter/ Instagram'])
                    program_name = get_field_value([
                        'Qualification/ Program Title',  # Primary column name in Excel
                        'Qualification/Program Title',
                        'Program Title',
                        'Qualification',
                        'program', 'Program', 
                        'program_name', 'Program Name',
                        'Course', 'Course Name',
                        'Program Profile'  # Fallback for section headers
                    ])
                    
                    # Get full name for display
                    full_name = self.get_full_name_display(first_name, last_name, middle_name)
                    
                    # Find or create program
                    program = self.find_or_create_program(program_name) if program_name else None
                    
                    # ===== CREATE/UPDATE LEARNER PROFILE (for Applicant Profile tab) =====
                    learner_profile = None
                    applicant = None
                    
                    # Try to find existing learner profile
                    if email:
                        learner_profile = Learner_Profile.objects.filter(email__iexact=email).first()
                    if not learner_profile and first_name and last_name:
                        learner_profile = Learner_Profile.objects.filter(
                            first_name__iexact=first_name,
                            last_name__iexact=last_name
                        ).first()
                    
                    if learner_profile:
                        # Update existing profile
                        learner_profile.entry_date = default_entry_date
                        if not learner_profile.email and email:
                            learner_profile.email = email
                        learner_profile.save()
                        total_profiles_updated += 1
                        if learner_profile.user:
                            applicant = learner_profile.user
                    else:
                        # Create new learner profile
                        # Provide defaults for required fields
                        learner_profile = Learner_Profile.objects.create(
                            last_name=last_name or 'Unknown',
                            first_name=first_name or 'Unknown',
                            middle_name=middle_name or '',
                            email=email or f'imported_{row_idx}@example.com',
                            entry_date=default_entry_date,
                            region='NCR',  # Default required field
                            province='Metro Manila',  # Default required field
                            city='Manila',  # Default required field
                            barangay='Unknown',  # Default required field
                            street='Unknown',  # Default required field
                            nationality='Filipino',
                        )
                        total_profiles_created += 1
                    
                    # ===== CREATE/UPDATE LMSTC_DOCUMENT (for Document Search tab) =====
                    # Create document name
                    document_name = f"{full_name} - Applicant Profile"
                    
                    # Create description
                    description_parts = []
                    if program_name:
                        description_parts.append(f"Program: {program_name}")
                    description_parts.append(f"Batch: {batch}")
                    description_parts.append(f"Year: 2024")
                    description = "; ".join(description_parts)
                    
                    # Check if document already exists
                    existing_doc = None
                    if learner_profile:
                        existing_doc = LMSTC_Documents.objects.filter(
                            learner_profile=learner_profile,
                            document_type='applicant_profile',
                            batch=batch,
                            uploaded_at__year=2024
                        ).first()
                    
                    if existing_doc:
                        # Update existing document
                        existing_doc.document_name = document_name
                        existing_doc.description = description
                        existing_doc.program = program
                        existing_doc.batch = batch
                        existing_doc.uploaded_at = default_upload_date
                        existing_doc.save()
                        total_docs_updated += 1
                    else:
                        # Create new document record
                        document = LMSTC_Documents.objects.create(
                            document_name=document_name,
                            document_type='applicant_profile',
                            description=description,
                            learner_profile=learner_profile,
                            applicant=applicant,
                            program=program,
                            batch=batch,
                            uploaded_by=user,
                            uploaded_at=default_upload_date,
                            status='active',
                            file_size=0,
                            mime_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                        total_docs_created += 1
                    
                except Exception as row_error:
                    error_msg = f"Sheet '{sheet_name}', Row {row_idx}: {str(row_error)}"
                    errors.append(error_msg)
                    total_skipped += 1
                    if len(errors) <= 10:  # Only show first 10 errors
                        self.stdout.write(self.style.WARNING(f'  Error processing row {row_idx}: {str(row_error)}'))
                    continue
        
        # Print summary
        self.stdout.write(self.style.SUCCESS(f'\n{"="*60}'))
        self.stdout.write(self.style.SUCCESS('Import completed!'))
        self.stdout.write(f'\nDocument Search (LMSTC_Documents):')
        self.stdout.write(f'  Created: {total_docs_created} records')
        self.stdout.write(f'  Updated: {total_docs_updated} records')
        self.stdout.write(f'\nApplicant Profile (Learner_Profile):')
        self.stdout.write(f'  Created: {total_profiles_created} records')
        self.stdout.write(f'  Updated: {total_profiles_updated} records')
        self.stdout.write(f'\nSkipped: {total_skipped} rows')
        
        if errors:
            self.stdout.write(self.style.WARNING(f'\nTotal errors: {len(errors)}'))
            if len(errors) > 10:
                self.stdout.write(self.style.WARNING(f'(Showing first 10 errors)'))
                for error in errors[:10]:
                    self.stdout.write(self.style.WARNING(f'  - {error}'))




